import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
from cmasher import get_sub_cmap
import matplotlib.colors as colors
from scipy.signal import peak_widths
from scipy.spatial.distance import directed_hausdorff
import sys


def read_data(data_file, correspondence_table_file):
    """Function to read data from measurement .bin file"""
    # .bin measurements file
    dt = np.dtype("uint16")
    f = open(data_file, "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)

    # correspondence of QDC number and strip number file
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_events = np.size(zdata) // 309

    # time conversion in seconds (integration time = 0.01s + 0.0005 s of dead time)
    time_values = [event * 0.0105 for event in range(nb_events)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, len(time_values)))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_events):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )

    return time_values, raw_strip_resp


def get_excel_data(excel_file, ws_name, start_l, start_col, len_data):
    """gets  noize and normalization values from excel file calculated from lateral
    microbeam scan"""
    wb_res = load_workbook(excel_file)
    ws = wb_res[ws_name]
    excel_data = [
        ws.cell(row=start_l + i, column=start_col).value for i in range(len_data)
    ]
    np.concatenate((np.zeros(18), excel_data))
    return np.array(excel_data)


def get_and_format_noize_normal_val(
    excel_file, ws_name, start_l, start_col, time_values
):
    noize = get_excel_data(excel_file, ws_name, start_l, start_col, 135)
    normal_val = get_excel_data(excel_file, ws_name, start_l, start_col + 1, 135)
    noize = np.repeat(noize, len(time_values)).reshape(-1, len(time_values))
    normal_val = np.repeat(normal_val, len(time_values)).reshape(-1, len(time_values))

    return noize, normal_val


def find_fwhm_values(peak_values):
    """Function returning the left and right half-maximum peak_values of a peak"""
    # Convert to numpy array for easier manipulation
    peak_values = np.array(peak_values)

    # Find the index of the peak
    peak_index = np.argmax(peak_values)
    peak_value = peak_values[peak_index]

    # Calculate the half maximum
    half_max = peak_value / 2.0

    # Find the left half-maximum point
    left_index = np.where(peak_values[:peak_index] <= half_max)[0][-1]
    right_index = np.where(peak_values[peak_index:] <= half_max)[0][0] + peak_index

    left_hm = peak_values[left_index]
    right_hm = peak_values[right_index]

    return left_hm, right_hm


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def find_x_closest_peak(x_data, y_data, x_target):
    x_peaks_index, _ = find_peaks(y_data)
    x_peaks = x_data[x_peaks_index]
    closest_x_peak = x_peaks[np.abs(x_peaks - x_target).argmin()]
    return closest_x_peak


def hausdorff_distance(set1, set2):
    """
    Calculate the Hausdorff distance between two sets of points.

    Parameters:
    set1 (numpy.ndarray): First set of points, shape (n_points1, 2)
    set2 (numpy.ndarray): Second set of points, shape (n_points2, 2)

    Returns:
    float: Hausdorff distance between set1 and set2
    """
    # Compute the directed Hausdorff distances
    d1 = directed_hausdorff(set1, set2)[0]
    d2 = directed_hausdorff(set2, set1)[0]

    # The Hausdorff distance is the maximum of the two directed distances
    return max(d1, d2)


def get_meas_cont(points):
    """Function to get the measured contours by the detector. Measured contours
    correspond to the FWHM values of each microbeam y profiles"""

    # gets miccrobeams x position
    x_mb_coord = [x for [x, _, v] in points if v > 40]

    # suppress every repeated x coord in x_mb_coord list
    x_mb_coord = list(set(x_mb_coord))

    # gets microbeam profile in Y
    mb_v = []
    mb_y = []
    for x_val in x_mb_coord:
        # corresponding y and v peak_values of microbeam considered
        mb_v_for_x_val = points[points[:, 0] == x_val][:, 2]
        mb_y_for_x_val = points[points[:, 0] == x_val][:, 1]

        # store v and peak_values of microbeams in a list
        mb_v.append(mb_v_for_x_val)
        mb_y.append(mb_y_for_x_val)

    # gets the half values of the microbeam FWHM
    fwhm_values = [find_fwhm_values(mb) for mb in mb_v]

    # gets the y coordinates of the half values of the microbeam FWHM
    y_mb_coord = []

    for index_mb, fwhm_tuple in enumerate(fwhm_values):
        # gets index of first half value in mb_v list
        index_y_half_val_1 = np.where(mb_v[index_mb] == fwhm_tuple[0])

        # gets y value corresponding to the index in mb_y list and adds it to y_mb_coord list
        y_mb_coord.append(mb_y[index_mb][index_y_half_val_1])

        # does the same for the second half value
        index_y_half_val_2 = np.where(mb_v[index_mb] == fwhm_tuple[1])
        y_mb_coord.append(mb_y[index_mb][index_y_half_val_2])

    y_mb_coord = np.array(y_mb_coord)
    # repeats every x coordinates of x_mb_coord twice (because for two y coord there is one
    # x coord)
    x_mb_coord_double = np.array([x for x in x_mb_coord for _ in range(2)])

    return np.column_stack((x_mb_coord_double, y_mb_coord))


def find_central_strip(points, dic_strip_pos, plot_x_profile=False):
    """Function finding the mid y of measurements and calc an X profile at this y
    position. Finds the first and last peak of this profile and calculates the middle
    x pos between these 2 peaks. Returns the strip number corresponding to this x pos.
    If plot_x_profile = True, plots the X profile with the first and last peak and the
    central strip."""

    # calc X profile at middle of Y axis
    theo_y_mid = (max(points[:, 1]) - min(points[:, 1])) / 2
    mid_y = y[find_closest_index(points[:, 1], theo_y_mid)]
    x_profile_x = points[np.where(points[:, 1] == mid_y)][:, 0]
    x_profile_v = points[np.where(points[:, 1] == mid_y)][:, 2]

    # find x of first and last peak of x profile
    peaks, _ = find_peaks(x_profile_v, height=20)
    x_peak_1 = x_profile_x[peaks[0]]
    x_peak_2 = x_profile_x[peaks[-1]]

    # find strip x with closest x pos to middle of distance between first and last peak
    x_target = abs(x_peak_2 - x_peak_1) / 2 + x_peak_2
    min_diff = np.abs(x_profile_x - x_target).min()
    central_strip_x = x_profile_x[np.abs(x_profile_x - x_target) == min_diff]
    central_strip = next(
        (k for k, v in dic_strip_pos.items() if v == central_strip_x), None
    )

    if plot_x_profile:
        plt.plot(x_profile_x, x_profile_v)
        plt.scatter(x_peak_1, x_profile_v[peaks[0]], color="y", label="First peak")
        plt.scatter(x_peak_2, x_profile_v[peaks[-1]], color="b", label="Last peak")
        plt.scatter(
            central_strip_x,
            x_profile_v[np.where(x_profile_x == central_strip_x)[0][0]],
            color="r",
            label="Central strip",
        )
        plt.xlabel(
            "Distance between strips at mask position (mm)", fontsize=fontsize_value
        )
        plt.ylabel("Normalized response (%)", fontsize=fontsize_value)
        plt.title("X profile at middle of Y axis", fontsize=fontsize_value)
        plt.tick_params(axis="both", which="major", labelsize=fontsize_value)
        plt.legend(loc="best")
        plt.show()

    return central_strip


def get_y_coord_of_center_y_profile(couch_shift, strip_resp):
    """Fonction calculating the FWHM of a given strip Y profile and return the y coord
    of the center of the FWHM"""

    # calc FWHM middle strip
    max_index = strip_resp.argmax()
    fwhm = peak_widths(strip_resp, [max_index], rel_height=0.5)
    fwhm_left = np.interp(fwhm[2][0], np.arange(len(couch_shift)), couch_shift)
    fwhm_right = np.interp(fwhm[3][0], np.arange(len(couch_shift)), couch_shift)

    return fwhm_right, fwhm_left, (fwhm_right - fwhm_left) / 2 + fwhm_left


def find_y_center_of_central_strip(
    strip_resp, central_strip, couch_shift, plot_y_profiles
):
    """Function calc the center of y profiles FWHM of strips neighboring central strip
    (because central strip is not in front of a microbeam) and returns average value"""

    # calc mid of y profile of strips neighboring central strip
    strip_neigh_1 = strip_resp[central_strip - 1, :]
    strip_neigh_2 = strip_resp[central_strip + 1, :]
    fwhm_l_1, fwhm_r_1, y_center_y_profile_1 = get_y_coord_of_center_y_profile(
        couch_shift, strip_neigh_1
    )
    fwhm_l_2, fwhm_r_2, y_center_y_profile_2 = get_y_coord_of_center_y_profile(
        couch_shift, strip_neigh_2
    )

    if plot_y_profiles:
        plt.plot(
            couch_shift,
            strip_neigh_2,
            label="y profile strip left to central strip",
        )
        plt.scatter(
            y_center_y_profile_2,
            max(strip_neigh_2),
            color="r",
        )
        plt.axvline(fwhm_l_2, color="r", linestyle="--", label="FWHM left")
        plt.axvline(fwhm_r_2, color="r", linestyle="--", label="FWHM right")
        plt.xlim(-5, 25)
        plt.xlabel("Couch shift (mm)", fontsize=fontsize_value)
        plt.ylabel("Normalized response (%)", fontsize=fontsize_value)
        plt.tick_params(axis="both", which="major", labelsize=fontsize_value)
        plt.show()

    return (y_center_y_profile_2 + y_center_y_profile_1) / 2


def find_closest_coordinates(array1, array2):
    """Function to find the closest coordinates in array2 to each coordinate in array1"""
    # Initialize an empty list to store the closest coordinates
    closest_coords = []
    # Loop through each coordinate pair in the first array
    for coord1 in array1:
        # Calculate the Euclidean distance to all coordinates in the second array
        distances = np.linalg.norm(array2 - coord1, axis=1)

        # Find the index of the closest coordinate in array2
        min_index = np.argmin(distances)

        # Get the closest coordinate pair from array2
        closest_coords.append(array2[min_index])

    # Convert the result to a numpy array and return
    return np.array(closest_coords)


## TO FILL #######
plot_raw_resp = False
plot_strip_resp = False
plot_x_profile = False
plot_y_profiles = False
fontsize_value = 15
CORRESPONDANCE_FILE = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
DATA_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\fx_formes_complexes\fx_circulaire\mesure\zData_150V_150ubeam_795mu_24p8v0_40collim17mmvitesse10.bin"
DCM_FILE = "RP.1.3.6.1.4.1.33868.20201021131037.169743"
COUCH_SPEED = 10
THRESHOLD = 0.15
NORMAL_VAL_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"

# strip exact positioning
strip_pos_file = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\microbeam_scan_analysis\strip_exact_positioning.xlsx"
strip_pos = get_excel_data(strip_pos_file, "strip_pos", 5, 7, 153)

# strip positionning at mask
pitch_strip = 0.2325
pitch_at_mask = 0.2075
dim_fact = pitch_at_mask / pitch_strip
strip_pos_at_mask = strip_pos * dim_fact
dic_strip_pos = {i: strip_pos_at_mask[i] for i in range(153)}

time_values, raw_strip_resp = read_data(DATA_FILE, CORRESPONDANCE_FILE)
couch_shift = np.array(time_values) * COUCH_SPEED
noize_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4, 135)
normal_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 5, 135)


# cut noize and normalize strip resp
noize_val, normal_val = get_and_format_noize_normal_val(
    NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4, time_values
)
strip_resp = (raw_strip_resp[18:153] - noize_val) / normal_val * 100
strip_resp = np.concatenate((np.zeros((18, len(time_values))), strip_resp))


# retrieves measurements whose resp > threshold and stores them in "points" array
points = np.empty((0, 3))
for istrip, strip in enumerate(strip_resp):
    for iresp, resp in enumerate(strip):
        if resp >= THRESHOLD:
            couch_mvt = time_values[iresp] * COUCH_SPEED
            # dist_between_strips = istrip * STRIP_PITCH
            points = np.append(
                points, [(dic_strip_pos[istrip], couch_mvt, resp)], axis=0
            )
x, y, v = points[:, 0], points[:, 1], points[:, 2]

# starts Y axis at 0
min_y = min(y)
y = y - min_y
points[:, 1] = points[:, 1] - min_y
couch_shift = couch_shift - min_y

# find central strip and its x pos = x pos of theo circle beam shape
central_strip = find_central_strip(points, dic_strip_pos, plot_x_profile)
x_center_circ = dic_strip_pos[central_strip]

# find y center of central strip = y pos of theo circle beam shape
y_center_circ = find_y_center_of_central_strip(
    strip_resp, central_strip, couch_shift, plot_y_profiles
)

# get measured contours
meas_cont = get_meas_cont(points)

# create theoratical circle shape of 17 mm centered at calculated points with a lot of
# points
rad = 17 / 2
theta = np.linspace(0, 2 * np.pi, 100000)
x_circ = x_center_circ + rad * np.cos(theta)
y_circ = y_center_circ + rad * np.sin(theta)
theo_circ_cont = np.column_stack((x_circ, y_circ))

# for each meas cont, find the closest theo point on theo shape (with a lot of points)
# to achieve same sampling for Hausdorff distance calc
circ_cont = find_closest_coordinates(meas_cont, theo_circ_cont)


# Plot theoratical and measured beam shapes
fig, ax = plt.subplots()
cmap = get_sub_cmap("YlGn", 0.2, 0.8)
norm = colors.Normalize(vmin=np.min(v), vmax=np.max(v))

# ax.plot(x_circ, y_circ, "k-", label="17 mm circle beam shape")
im = ax.scatter(x, y, s=30, marker="s", cmap=cmap, vmin=v.min(), vmax=v.max(), c=v)
ax.scatter(
    meas_cont[:, 0],
    meas_cont[:, 1],
    s=30,
    marker="s",
    color="red",
    label="measured contours",
)
ax.scatter(circ_cont[:, 0], circ_cont[:, 1], label="theoratical contours")
cbar = fig.colorbar(im, label="Normalized response (%)")
cbar.ax.tick_params(labelsize=fontsize_value)
cbar.set_label("Normalized response (%)", fontsize=fontsize_value)
ax.set_xlabel("Distance between strips at mask position (mm)", fontsize=fontsize_value)

# To get same scale on both axis
x_ticks = np.arange(-12, 14, 2)
y_ticks = np.arange(-1, 23, 2)
ax.set_xticks(x_ticks)
ax.set_yticks(y_ticks)
ax.set_aspect("equal", adjustable="box")
ax.set_xlabel("Distance (mm)", fontsize=fontsize_value)
ax.set_ylabel("Distance (mm)", fontsize=fontsize_value)
ax.tick_params(axis="both", which="major", labelsize=fontsize_value)
plt.legend(fontsize=fontsize_value)
plt.show()

# Calc Hausdorff distance
distance = hausdorff_distance(circ_cont, meas_cont)
print(f"Hausdorff distance: {distance}")
