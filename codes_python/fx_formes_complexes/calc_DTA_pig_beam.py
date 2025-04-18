import matplotlib.pyplot as plt
import matplotlib.colors as colors
import numpy as np
import math
import pydicom
import sys
from openpyxl import load_workbook
from scipy.interpolate import interp1d
from cmasher import get_sub_cmap
from scipy.signal import find_peaks, peak_widths
from scipy.spatial.distance import directed_hausdorff


def read_data(data_file, correspondence_table_file):
    """Function to read data from measurement .bin file"""
    # .bin measurements file
    dt = np.dtype("uint16")
    f = open(data_file, "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)

    # correspondence of QDC number and strip number file
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_events = np.size(zdata) // 309

    # time conversion in seconds (integration time = 0.01s + 0.0005 s of dead time)
    time_values = [event * 0.0105 for event in range(nb_events)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, len(time_values)))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_events):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )

    return time_values, raw_strip_resp


def mask_hull(tuple_points):
    """Function to compute the convex hull of a set of points"""
    points = []
    for [x, y] in tuple_points:
        middle = math.ceil(len(points) / 2)
        if len(points) >= 2 and points[middle - 1][0] == x and points[middle][0] == x:
            points.pop(middle)
        points.insert(middle, (x, y))

    return points


def get_beam_contour_from_dcm(plan_name):
    """Function to read dcm file containing beam shape (from TPS) and returning beam
    contours"""
    plan = pydicom.read_file(plan_name, force=True)
    dummy_X, dummy_Y, centroid = [], [], []
    beam_seq = plan[(0x300A, 0x00B0)].value
    beam_contours = []
    for i in range(len(beam_seq)):
        block_seq = beam_seq[i][(0x300A, 0x00F4)].value
        block_data = block_seq[0][(0x300A, 0x0106)].value
        dummy = []
        for entry in block_data:
            dummy.append(entry)
        dummy_array = np.array(dummy).astype(float)
        dummy_X.append(dummy_array[::2])
        dummy_X[i] = dummy_X[i].reshape(dummy_X[i].shape[0], 1)
        dummy_Y.append(dummy_array[1::2])
        dummy_Y[i] = dummy_Y[i].reshape(dummy_Y[i].shape[0], 1)
        dummy_X = dummy_X
        dummy_Y = dummy_Y
        centroid.append(
            (
                sum(dummy_X[i][:, 0]) / len(dummy_X[i][:, 0]),
                sum(dummy_Y[i][:, 0]) / len(dummy_Y[i][:, 0]),
            )
        )
        beam_contours.append(np.concatenate((dummy_X[i], dummy_Y[i]), axis=1))
    return beam_contours


def get_excel_data(excel_file, ws_name, start_l, start_col):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(excel_file)
    ws = wb_res[ws_name]
    excel_data = [ws.cell(row=start_l + i, column=start_col).value for i in range(135)]
    np.concatenate((np.zeros(18), excel_data))
    return np.array(excel_data)


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def get_and_format_noize_normal_val(
    excel_file, ws_name, start_l, start_col, time_values
):
    noize = get_excel_data(excel_file, ws_name, start_l, start_col)
    normal_val = get_excel_data(excel_file, ws_name, start_l, start_col + 1)
    noize = np.repeat(noize, len(time_values)).reshape(-1, len(time_values))
    noize_strips = np.concatenate((np.zeros((18, len(time_values))), noize))
    normal_val = np.repeat(normal_val, len(time_values)).reshape(-1, len(time_values))
    normal_val_strips = np.concatenate((np.zeros((18, len(time_values))), normal_val))

    return noize_strips, normal_val_strips


def plot_resp(time, resp, title, y_label):
    for strip_num in range(18, 153):
        plt.plot(time, resp[strip_num, :], label=f"Strip {strip_num}")
    plt.xlabel("Time (s)")
    plt.ylabel(y_label)
    plt.legend(loc="best")
    plt.title(title, fontsize=18, fontstyle="oblique")
    plt.show()


def get_central_x_val_fwhm(x, y):
    """Function to return the central value of the FWHM of a profile"""

    max_index = y.argmax()
    fwhm = peak_widths(y, [max_index], rel_height=0.5)
    fwhm_left = np.interp(fwhm[2][0], np.arange(len(x)), x)
    fwhm_right = np.interp(fwhm[3][0], np.arange(len(x)), x)

    return (fwhm_right - fwhm_left) / 2 + fwhm_left


def cut_beam_contours_at_given_x(beam_contours, x_cut):
    """Function returning beam_contours cut at x_cut"""

    beam_contours_cut = np.column_stack(
        (
            beam_contours[beam_contours[:, 0] > x_cut][:, 0],
            beam_contours[beam_contours[:, 0] > x_cut][:, 1],
        )
    )

    # cut beam contours in 2 : one top and one bottom part to interpolate separately edges
    # of cut shape, thresh y = 10 mm
    beam_contours_top = np.column_stack(
        (
            beam_contours[beam_contours[:, 1] > 10][:, 0],
            beam_contours[beam_contours[:, 1] > 10][:, 1],
        )
    )
    beam_contours_bott = np.column_stack(
        (
            beam_contours[beam_contours[:, 1] < 10][:, 0],
            beam_contours[beam_contours[:, 1] < 10][:, 1],
        )
    )

    # get top edge contour
    top_interpolation = interp1d(
        beam_contours_top[:, 0],
        beam_contours_top[:, 1],
        kind="linear",
        fill_value="extrapolate",
    )
    y_cut_top = top_interpolation(x_cut)

    # adds top edge contour to arrays of the cut shape coordinates
    # beam_contours_cut = np.append(beam_contours_cut, [[x_cut, y_cut_top]], axis=0)

    # get bottom edge contour
    bott_interpolation = interp1d(
        beam_contours_bott[:, 0], beam_contours_bott[:, 1], kind="linear"
    )
    y_cut_bott = bott_interpolation(x_cut)

    # adds top edge contour to arrays of the cut shape coordinates
    beam_contours_cut = np.append(beam_contours_cut, [[x_cut, y_cut_bott]], axis=0)

    return beam_contours_cut


def register_beam_contours(
    beam_contours,
    extreme_right_strip,
    STRIP_PITCH,
    strip_width,
    inter_strip_width,
    strip_resp,
    couch_shift,
):
    """Function registering the beam contours from TPS to measurements on the right
    side of the pig beam by doing a translation in x and y direction"""

    # right side x position of mask shape
    x_r_side_mask = (
        extreme_right_strip * STRIP_PITCH + strip_width / 2 + inter_strip_width
    )

    # translate mask shape to match measurements in x direction
    x_shift = abs(np.max(beam_contours[:, 0]) - x_r_side_mask)
    beam_contours[:, 0] = beam_contours[:, 0] + x_shift

    # Strip furthest on the right of all the strips facing a microbeam
    max_strip_resp = np.max(strip_resp, axis=1)
    r_mb_strip = (np.where(max_strip_resp > 90)[0])[-1]

    # get central value of FWHM of the profile of the furthest right strip = y center
    # position of the .dcm beam shape
    y_center_r_mb_strip = get_central_x_val_fwhm(couch_shift, strip_resp[r_mb_strip, :])

    # get max x coord of beam contours
    max_x_dcm = np.max(beam_contours[:, 0])

    # get y dcm coord of furthest right side of TPS beam shape
    y_r_side_coord = beam_contours[beam_contours[:, 0] == max_x_dcm][:, 1]

    # compare center of right side TPS beam shame and center right mb strip = y shift
    y_r_side_center = (y_r_side_coord[1] - y_r_side_coord[0]) / 2 + y_r_side_coord[0]
    y_shift = abs(y_r_side_center - y_center_r_mb_strip)
    beam_contours[:, 1] = beam_contours[:, 1] + y_shift

    return beam_contours


def find_fwhm_values(peak_values):
    """Function returning the left and right half-maximum peak_values of a peak"""
    # Convert to numpy array for easier manipulation
    peak_values = np.array(peak_values)

    # Find the index of the peak
    peak_index = np.argmax(peak_values)
    peak_value = peak_values[peak_index]

    # Calculate the half maximum
    half_max = peak_value / 2.0

    # Find the left half-maximum point
    left_index = np.where(peak_values[:peak_index] <= half_max)[0][-1]
    right_index = np.where(peak_values[peak_index:] <= half_max)[0][0] + peak_index

    left_hm = peak_values[left_index]
    right_hm = peak_values[right_index]

    return left_hm, right_hm


def hausdorff_distance(set1, set2):
    """
    Calculate the Hausdorff distance between two sets of points.

    Parameters:
    set1 (numpy.ndarray): First set of points, shape (n_points1, 2)
    set2 (numpy.ndarray): Second set of points, shape (n_points2, 2)

    Returns:
    float: Hausdorff distance between set1 and set2
    """
    # Compute the directed Hausdorff distances
    d1 = directed_hausdorff(set1, set2)[0]
    d2 = directed_hausdorff(set2, set1)[0]

    # The Hausdorff distance is the maximum of the two directed distances
    return max(d1, d2)


def calc_gamma_dta(coords_ref, coords_eval, dta_thresh):
    """
    Compare 2 sets of coordinates with a criteria Distance To Agreement (DTA).

    Args:
        ref_coordinates (array-like): list of de tuples (x, y) for reference coordinates.
        eval_coordinates (array-like): list of de tuples (x, y) for coordinates to evaluate.
        dta_threshold (float): distance threshold for DTA criteria.

    Returns:
        float: mean gamma index for 2 sets of coordinates.
    """

    ref_coordinates = np.array(coords_ref)
    eval_coordinates = np.array(coords_eval)

    gamma_indexes = []

    # compare each point from eval_coordinates with each point from ref_coordinates
    for eval_coord in eval_coordinates:
        # set min dist to - infinity
        min_distance = np.inf
        for ref_coord in ref_coordinates:
            # calc Euclidean distance between 2 coordinates
            distance = np.linalg.norm(eval_coord - ref_coord)

            # Find min dist
            if distance < min_distance:
                min_distance = distance

        # calc gamma index for this point
        gamma_index = min_distance / dta_thresh
        gamma_indexes.append(gamma_index)

    # Calculer la valeur moyenne du gamma index
    mean_gamma_index = np.mean(gamma_indexes)

    return mean_gamma_index


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l + i, column=start_col, value=data_to_fill[i])
    wb_res.save(excel_path)


# TO FILL #############
plot_raw_resp = False
plot_strip_resp = False
fontsize_value = 10
fill_excel_bool = False
# strip measuring the microbeam on the furthest right of the mask
extreme_right_strip = 146
excel_path = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\fx_formes_complexes\calc_gamma_index.xlsx"
ws_name = "pig_beam_shape"
start_l = 4
start_c = 9

CORRESPONDANCE_FILE = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
DATA_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\fx_formes_complexes\fx_cochons\mesures\zData_150V_150ubeam_795mu_24p8v0_40_collimCochon_vitesse10.bin"
DCM_FILE = r"C:\Users\milewski\Desktop\these\phd\codes_python\fx_formes_complexes\RP.1.3.6.1.4.1.33868.20210210162030.730598"
NORMAL_VAL_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"

# strip pitch 0.2075 mm calculated at mask position
STRIP_PITCH = 0.2075
COUCH_SPEED = 10
THRESHOLD = 0.15


time_values, raw_strip_resp = read_data(DATA_FILE, CORRESPONDANCE_FILE)
couch_shift = np.array(time_values) * COUCH_SPEED
noize_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4)
normal_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 5)
strip_width = 0.1725  # mm
inter_strip_width = 0.06  # mm


# raw strips response plot
if plot_raw_resp:
    plot_resp(
        time_values,
        raw_strip_resp,
        "Raw responses pig beam ESRF",
        "QDC response (AU)",
    )


# cut noize and normalize strip resp
noize_val, normal_val = get_and_format_noize_normal_val(
    NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4, time_values
)
strip_resp = (raw_strip_resp - noize_val) / normal_val * 100

# strips response plot
if plot_strip_resp:
    plot_resp(
        time_values,
        strip_resp,
        "Normalized strip responses pig beam ESRF",
        "Normalized resp (%)",
    )


# Compute points
points = np.empty((0, 3))
for istrip, strip in enumerate(strip_resp):
    for iresp, resp in enumerate(strip):
        if resp >= THRESHOLD:
            dist_between_strips = istrip * STRIP_PITCH
            points = np.append(
                points, [(dist_between_strips, couch_shift[iresp], resp)], axis=0
            )

x, y, v = points[:, 0], points[:, 1], points[:, 2]

# to start Y axis at 0
min_y = min(y)
y = [y[i] - min_y for i in range(len(y))]
points[:, 1] = points[:, 1] - min_y
couch_shift = couch_shift - min_y

# beam contours from dcm
beam_cont_dcm = get_beam_contour_from_dcm(DCM_FILE)
beam_contours = beam_cont_dcm[1][:-1]

# add first point of beam contours at the end to complete the shape
beam_contours = np.vstack([beam_contours, beam_contours[0]])

# reverse beam contours shape to retrieve same orientation as strip measurements
beam_contours[:, 1] = -beam_contours[:, 1]

# register beam contours to strip measurements
beam_contours = register_beam_contours(
    beam_contours,
    extreme_right_strip,
    STRIP_PITCH,
    strip_width,
    inter_strip_width,
    strip_resp,
    couch_shift,
)

# x coord swhere to cut TPS beam shape to compare shapes with Hausdorff distance
x_cut = 18 * STRIP_PITCH - strip_width / 2

# cut beam contours at x_cut
beam_contours_cut = cut_beam_contours_at_given_x(beam_contours, x_cut)

# gets miccrobeams x position
x_mb_coord = [x for [x, _, v] in points if v > 40]

# suppress every repeated x coord in x_mb_coord list
x_mb_coord = list(set(x_mb_coord))

# gets microbeam profile in Y
mb_v = []
mb_y = []
for x_val in x_mb_coord:
    # corresponding y and v peak_values of microbeam considered
    mb_v_for_x_val = points[points[:, 0] == x_val][:, 2]
    mb_y_for_x_val = points[points[:, 0] == x_val][:, 1]

    # store v and peak_values of microbeams in a list
    mb_v.append(mb_v_for_x_val)
    mb_y.append(mb_y_for_x_val)

# gets the half values of the microbeam FWHM
fwhm_values = [find_fwhm_values(mb) for mb in mb_v]

# gets the y coordinates of the half values of the microbeam FWHM
y_mb_coord = []


for index_mb, fwhm_tuple in enumerate(fwhm_values):
    # gets index of first half value in mb_v list
    index_y_half_val_1 = np.where(mb_v[index_mb] == fwhm_tuple[0])

    # gets y value corresponding to the index in mb_y list and adds it to y_mb_coord list
    y_mb_coord.append(mb_y[index_mb][index_y_half_val_1])

    # does the same for the second half value
    index_y_half_val_2 = np.where(mb_v[index_mb] == fwhm_tuple[1])
    y_mb_coord.append(mb_y[index_mb][index_y_half_val_2])

y_mb_coord = np.array(y_mb_coord)
# repeats every x coordinates of x_mb_coord twice (because for two y coord there is one
# x coord)
x_mb_coord_double = np.array([x for x in x_mb_coord for _ in range(2)])


# Plot beam shape
fontsize_value = 15
fig, ax = plt.subplots()
ax.scatter(beam_contours_cut[:, 0], beam_contours_cut[:, 1])
ax.scatter(x_mb_coord_double, y_mb_coord, s=30, marker="s", color="red")
ax.set_xlabel("Distance (mm)")
ax.set_ylabel("Distance (mm)")

cmap = get_sub_cmap("YlGn", 0.2, 0.8)
norm = colors.Normalize(vmin=np.min(v), vmax=np.max(v))
im = ax.scatter(x, y, s=5, cmap=cmap, vmin=v.min(), vmax=v.max(), c=v)
cbar = fig.colorbar(im, label="Normalized response (%)")
cbar.ax.tick_params(labelsize=fontsize_value)
cbar.set_label("Normalized response (%)", fontsize=fontsize_value)
ax.set_xlabel("Distance between strips at mask position (mm)", fontsize=fontsize_value)
ax.legend(fontsize=fontsize_value)

# To get same scale on both axis
ticks = np.arange(0, 32, 2)
ax.set_xticks(ticks)
ax.set_yticks(ticks)
ax.set_aspect("equal", adjustable="box")
# ax.set_ylabel("Couch shift (mm)", fontsize=fontsize_value)
# ax.set_ylim(-2, 24)
# ax.set_xlim(0, 34)
plt.show()


# Reorganize coordinates of theoratical shape and measured shape to have a list of
# tuples (x, y)
theo_coord = np.column_stack((beam_contours_cut[:, 0], beam_contours_cut[:, 1]))
mes_coord = np.column_stack((x_mb_coord_double, y_mb_coord))

# calc gamma index (mes_coord = ref_coordinates because contours from TPS are under
# sampled and if comparison is performed in the other way, the gamma index will be
# artificially high) for different x and y shift (to optimize registration between
# theoratical and measured shapes)
dta_threshold = 0.17
shift_pattern = np.concatenate(
    (np.arange(0.01, 0.06, 0.01), np.arange(-0.01, -0.06, -0.01))
)
x_shift = np.concatenate(
    (
        np.array([0]),
        shift_pattern,
        np.zeros(10),
        shift_pattern,
        np.arange(0.01, 0.06, 0.01),
        np.arange(-0.01, -0.06, -0.01),
    )
)
y_shift = np.concatenate(
    (
        np.zeros(11),
        shift_pattern,
        shift_pattern,
        np.arange(-0.01, -0.06, -0.01),
        np.arange(0.01, 0.06, 0.01),
    )
)

x_shift = np.arange(0.11, 0.21, 0.01)
y_shift = x_shift

x_mes, y_mes = mes_coord[:, 0], mes_coord[:, 1]

mean_gamma = []

for x_s, y_s in zip(x_shift, y_shift):
    x_mes_shifted = x_mes + x_s
    y_mes_shifted = y_mes + y_s
    mean_gamma.append(
        calc_gamma_dta(
            np.column_stack((x_mes_shifted, y_mes_shifted)),
            theo_coord,
            dta_threshold,
        )
    )

# Plot beam shape
fontsize_value = 15
fig, ax = plt.subplots()
ax.scatter(mes_coord[:, 0], mes_coord[:, 1], s=5, marker="s", color="yellow")
ax.scatter(x_mes_shifted, y_mes_shifted, s=5, marker="s", color="red")
ax.scatter(theo_coord[:, 0], theo_coord[:, 1], s=5, marker="s", color="blue")
ax.set_xlabel("Distance (mm)")
ax.set_ylabel("Distance (mm)")
plt.show()

if fill_excel_bool:
    fill_excel(excel_path, ws_name, x_shift, start_l, start_c)
    fill_excel(excel_path, ws_name, y_shift, start_l, start_c + 1)
    fill_excel(excel_path, ws_name, mean_gamma, start_l, start_c + 2)
