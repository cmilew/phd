import matplotlib.pyplot as plt
import matplotlib.colors as colors
import numpy as np
import math
from shapely.geometry import Polygon
import pydicom
import sys
from openpyxl import load_workbook
from cmasher import get_sub_cmap
from scipy.interpolate import interp1d
from scipy.signal import find_peaks


def read_data(data_file, correspondence_table_file):
    """Function to read data from measurement .bin file"""
    # .bin measurements file
    dt = np.dtype("uint16")
    f = open(data_file, "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)

    # correspondence of QDC number and strip number file
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_events = np.size(zdata) // 309

    # time conversion in seconds (integration time = 0.01s + 0.0005 s of dead time)
    time_values = [event * 0.0105 for event in range(nb_events)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, nb_events))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_events):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )

    return nb_events, time_values, raw_strip_resp


def mask_hull(tuple_points):
    """Function to compute the convex hull of a set of points"""
    points = []
    for [x, y] in tuple_points:
        middle = math.ceil(len(points) / 2)
        if len(points) >= 2 and points[middle - 1][0] == x and points[middle][0] == x:
            points.pop(middle)
        points.insert(middle, (x, y))

    return points


def get_beam_contour_from_dcm(plan_name):
    """Function to read dcm file containing beam shape (from TPS) and returning beam
    contours"""
    plan = pydicom.read_file(plan_name, force=True)
    dummy_X, dummy_Y, centroid = [], [], []
    beam_seq = plan[(0x300A, 0x00B0)].value
    beam_contours = []
    for i in range(len(beam_seq)):
        block_seq = beam_seq[i][(0x300A, 0x00F4)].value
        block_data = block_seq[0][(0x300A, 0x0106)].value
        dummy = []
        for entry in block_data:
            dummy.append(entry)
        dummy_array = np.array(dummy).astype(float)
        dummy_X.append(dummy_array[::2])
        dummy_X[i] = dummy_X[i].reshape(dummy_X[i].shape[0], 1)
        dummy_Y.append(dummy_array[1::2])
        dummy_Y[i] = dummy_Y[i].reshape(dummy_Y[i].shape[0], 1)
        dummy_X = dummy_X
        dummy_Y = dummy_Y
        centroid.append(
            (
                sum(dummy_X[i][:, 0]) / len(dummy_X[i][:, 0]),
                sum(dummy_Y[i][:, 0]) / len(dummy_Y[i][:, 0]),
            )
        )
        beam_contours.append(np.concatenate((dummy_X[i], dummy_Y[i]), axis=1))
    return beam_contours


def get_excel_data(excel_file, ws_name, start_l, start_col):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(excel_file)
    ws = wb_res[ws_name]
    excel_data = [ws.cell(row=start_l + i, column=start_col).value for i in range(135)]
    np.concatenate((np.zeros(18), excel_data))
    return np.array(excel_data)


def cut_noize_and_normalize(raw_strip_resp, strip_noize, normalization_value, factor):
    clean_data_fn = (
        lambda x: (x - strip_noize) / (normalization_value - strip_noize) * factor
    )
    vectorized_clean_data_fn = np.vectorize(clean_data_fn)
    normalized_strip_resp = vectorized_clean_data_fn(raw_strip_resp)
    return normalized_strip_resp


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l + i, column=start_col, value=data_to_fill[i])
    wb_res.save(excel_path)


def find_x_closest_peak(x_data, y_data, x_target):
    x_peaks_index, _ = find_peaks(y_data)
    x_peaks = x_data[x_peaks_index]
    closest_x_peak = x_peaks[np.abs(x_peaks - x_target).argmin()]
    return closest_x_peak


CORRESPONDANCE_FILE = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"

# circular collimator
DATA_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\fx_formes_complexes\fx_circulaire\mesure\zData_150V_150ubeam_795mu_24p8v0_40collim17mmvitesse10.bin"
DCM_FILE = "RP.1.3.6.1.4.1.33868.20201021131037.169743"
NORMAL_VAL_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"

# # raw strips response plot
# for strip_num in range(130, 153):
#     plt.plot(time_values, raw_strip_resp[strip_num, :], label=str(strip_num))
# plt.xlabel("Time (s)")
# plt.ylabel("QDC response (AU)")
# plt.legend(loc="best")
# plt.title(f"Raw responses pig beam ESRF", fontsize=18, fontstyle="oblique")
# plt.show()

# Initialize scatter plot data
THRESHOLD = 40
X_OFFSET = 500

# strip pitch 0.2075 mm calculated at mask position
STRIP_PITCH = 0.2075
COUCH_SPEED = 10

nb_events, time_values, raw_strip_resp = read_data(DATA_FILE, CORRESPONDANCE_FILE)
couch_shift = np.array(time_values) * COUCH_SPEED
normal_val = get_excel_data(NORMAL_VAL_FILE, "mes_ESRF", 4, 4)
noize_val = get_excel_data(NORMAL_VAL_FILE, "mes_ESRF", 4, 5)


# retrieves measurements whose resp > threshold and stores them in "points" array
points = np.empty((0, 3))
for istrip, strip in enumerate(raw_strip_resp):
    for iresp, resp in enumerate(strip):
        if resp >= THRESHOLD:
            couch_mvt = time_values[iresp] * COUCH_SPEED
            dist_between_strips = istrip * STRIP_PITCH
            points = np.append(points, [(dist_between_strips, couch_mvt, resp)], axis=0)

x, y, v = points[:, 0], points[:, 1], points[:, 2]

# starts Y axis at 0
min_y = min(y)
y = [y[i] - min_y for i in range(len(y))]
points[:, 1] = points[:, 1] - min_y

# get middle of y axis coordinates
theo_mid_y = (max(y) - min(y)) / 2
mid_y = y[find_closest_index(y, theo_mid_y)]

# x profile x and v values at middle of y axis
x_profil_x = points[np.in1d(points[:, 1], mid_y)][:, 0]
x_profil_v = points[np.in1d(points[:, 1], mid_y)][:, 2]

# integral of x profile at middle of y axis
integ_x_profile = np.zeros(len(x_profil_v))
for i in range(len(x_profil_v)):
    if i == 0:
        integ_x_profile[i] = x_profil_v[i] * STRIP_PITCH
    else:
        integ_x_profile[i] = x_profil_v[i] * STRIP_PITCH + integ_x_profile[i - 1]

# calc middle of integral => midle of beam field
mid_integ_y = (max(integ_x_profile) - min(integ_x_profile)) / 2
lin_interpol = interp1d(integ_x_profile, x_profil_x, kind="linear")
mid_integ_x = lin_interpol(mid_integ_y)

# find closest peak to middle of integral = x coord of theo circle center
x_center_circ = find_x_closest_peak(x_profil_x, x_profil_v, mid_integ_x)


# middle strip to use for Y profile
mid_strip = int(x_center_circ / STRIP_PITCH)
mid_strip_resp = raw_strip_resp[mid_strip, :]
mid_couch_shift = couch_shift - min_y

# sys.exit()
# FWHM middle strip
fwhm = np.max(mid_strip_resp) / 2
max_mid_resp_ind = int(np.where(mid_strip_resp == np.max(mid_strip_resp))[0])
up_lin_interp = interp1d(
    mid_strip_resp[0:max_mid_resp_ind],
    mid_couch_shift[0:max_mid_resp_ind],
    kind="linear",
)
fwhm_1 = up_lin_interp(fwhm)
down_lin_interp = interp1d(
    mid_strip_resp[max_mid_resp_ind:], mid_couch_shift[max_mid_resp_ind:], kind="linear"
)
fwhm_2 = down_lin_interp(fwhm)
y_center_circ = (fwhm_2 - fwhm_1) / 2 + fwhm_1

# create theoratical circle shape of 17 mm centered at calculated points
rad = 17 / 2
theta = np.linspace(0, 2 * np.pi, 100)
x_circ = x_center_circ + rad * np.cos(theta)
y_circ = y_center_circ + rad * np.sin(theta)

# Compute hull
hull_points = []
for [strip, resp, _] in points:
    # entre les pistes 69 et 117 on ne prend que les points sur les pistes impaires
    # le reste du temps on prends les points sur les pistes paires
    if (strip / STRIP_PITCH) % 2 == int(69 <= (strip / STRIP_PITCH) <= 117):
        hull_points.append((strip, resp))

hull = mask_hull(hull_points)
hx, hy = zip(*hull + [hull[0]])

# hull area
hull_area_points = []
for [strip, resp] in hull:
    hull_area_points.append((strip * STRIP_PITCH, resp * COUCH_SPEED))
hull_area = Polygon(hull_area_points).area


# Plot theoratical and measured beam shapes
fontsize_value = 15
fig, ax = plt.subplots()
ax.plot(x_circ, y_circ, "k-", label="Theoretical circle beam shape")
cmap = get_sub_cmap("YlGn", 0.2, 0.8)
norm = colors.Normalize(vmin=np.min(v), vmax=np.max(v))
im = ax.scatter(x, y, s=50, marker="s", cmap=cmap, vmin=v.min(), vmax=v.max(), c=v)
cbar = fig.colorbar(im, label="QDC response (AU)")
cbar.ax.tick_params(labelsize=fontsize_value)
cbar.set_label("QDC response (AU)", fontsize=fontsize_value)
ax.plot(hx, hy, "r-", label="Beam shape measured")
ax.set_xlabel("Distance between strips at mask position (mm)", fontsize=fontsize_value)
# To get same scale on both axis
ticks = np.arange(0, 30, 2)
ax.set_xticks(ticks)
ax.set_yticks(ticks)
ax.set_aspect("equal", adjustable="box")
ax.set_ylabel("Couch shift (mm)", fontsize=fontsize_value)
ax.set_ylim(-2, 22)
ax.set_xlim(0, 30)
plt.scatter(x_center_circ, y_center_circ, s=50, label="Center of circle shape")
plt.show()
