import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
from scipy.interpolate import interp1d
from cmasher import get_sub_cmap
import matplotlib.colors as colors
from scipy.signal import find_peaks, peak_widths
from scipy.spatial.distance import directed_hausdorff
import sys


def read_data(data_file, correspondence_table_file):
    """Function to read data from measurement .bin file"""
    # .bin measurements file
    dt = np.dtype("uint16")
    f = open(data_file, "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)

    # correspondence of QDC number and strip number file
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_events = np.size(zdata) // 309

    # time conversion in seconds (integration time = 0.01s + 0.0005 s of dead time)
    time_values = [event * 0.0105 for event in range(nb_events)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, len(time_values)))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_events):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )

    return time_values, raw_strip_resp


def get_excel_data(excel_file, ws_name, start_l, start_col):
    """gets  noize and normalization peak_values from excel file calculated from lateral
    microbeam scan"""
    wb_res = load_workbook(excel_file)
    ws = wb_res[ws_name]
    excel_data = [ws.cell(row=start_l + i, column=start_col).value for i in range(135)]
    np.concatenate((np.zeros(18), excel_data))
    return np.array(excel_data)


def get_and_format_noize_normal_val(
    excel_file, ws_name, start_l, start_col, time_values
):
    noize = get_excel_data(excel_file, ws_name, start_l, start_col)
    normal_val = get_excel_data(excel_file, ws_name, start_l, start_col + 1)
    noize = np.repeat(noize, len(time_values)).reshape(-1, len(time_values))
    normal_val = np.repeat(normal_val, len(time_values)).reshape(-1, len(time_values))

    return noize, normal_val


def find_fwhm_values(peak_values):
    """Function returning the left and right half-maximum peak_values of a peak"""
    # Convert to numpy array for easier manipulation
    peak_values = np.array(peak_values)

    # Find the index of the peak
    peak_index = np.argmax(peak_values)
    peak_value = peak_values[peak_index]

    # Calculate the half maximum
    half_max = peak_value / 2.0

    # Find the left half-maximum point
    left_index = np.where(peak_values[:peak_index] <= half_max)[0][-1]
    right_index = np.where(peak_values[peak_index:] <= half_max)[0][0] + peak_index

    left_hm = peak_values[left_index]
    right_hm = peak_values[right_index]

    return left_hm, right_hm


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def find_x_closest_peak(x_data, y_data, x_target):
    x_peaks_index, _ = find_peaks(y_data)
    x_peaks = x_data[x_peaks_index]
    closest_x_peak = x_peaks[np.abs(x_peaks - x_target).argmin()]
    return closest_x_peak


def get_y_coord_theo_circle(x_center_coord, couch_shift, strip_resp):
    # get mid strip resp
    mid_strip = int(x_center_coord / STRIP_PITCH)
    mid_strip_resp = strip_resp[mid_strip, :]

    # calc FWHM middle strip
    max_index = mid_strip_resp.argmax()
    fwhm = peak_widths(mid_strip_resp, [max_index], rel_height=0.5)
    fwhm_left = np.interp(fwhm[2][0], np.arange(len(couch_shift)), couch_shift)
    fwhm_right = np.interp(fwhm[3][0], np.arange(len(couch_shift)), couch_shift)

    return (fwhm_right - fwhm_left) / 2 + fwhm_left


def get_x_coord_theo_circle(points):
    """Function to calculate the x coordinate of the center of the theoretical beam
    circle shape. Integrates an X profile of measurements taken at middle of Y axis.
    Calculates abscisse of the middle of the integral = x coordinate of the center of
    theoratical circle"""

    # get middle of y axis coordinate
    theo_mid_y = (max(points[:, 1]) - min(points[:, 1])) / 2
    mid_y = y[find_closest_index(points[:, 1], theo_mid_y)]

    # x profile values (at middle of y axis)
    x_profil_x = points[np.in1d(points[:, 1], mid_y)][:, 0]
    x_profil_v = points[np.in1d(points[:, 1], mid_y)][:, 2]

    # integral of x profile
    integ = np.zeros(len(x_profil_v))
    for i in range(len(x_profil_v)):
        if i == 0:
            integ[i] = x_profil_v[i] * STRIP_PITCH
        else:
            integ[i] = x_profil_v[i] * STRIP_PITCH + integ[i - 1]

    # calc middle of integral
    mid_integ_y = (max(integ) - min(integ)) / 2
    mid_integ_x = np.interp(mid_integ_y, integ, x_profil_x)

    # find closest peak to middle of integral = x coord of theo circle center
    x_center_circ = find_x_closest_peak(x_profil_x, x_profil_v, mid_integ_x)

    return x_center_circ


def calc_dta(coords_ref, coords_eval, dta_thresh):
    """
    Compare 2 sets of coordinates with a criteria Distance To Agreement (DTA).

    Args:
        coords1 (array-like): Liste de tuples (x, y) pour la première série de coordonnées.
        coords2 (array-like): Liste de tuples (x, y) pour la deuxième série de coordonnées.
        dta_threshold (float): Le seuil de distance pour le critère DTA.

    Returns:
        float: Le gamma index moyen pour les deux séries de coordonnées.
    """
    ref_coordinates = np.array(coords_ref)
    eval_coordinates = np.array(coords_eval)

    gamma_indexes = []

    # compare each point from eval_coordinates with each point from ref_coordinates
    for eval_coord in eval_coordinates:
        # set min dist to - infinity
        min_distance = np.inf
        for ref_coord in ref_coordinates:
            # calc Euclidean distance between 2 coordinates
            distance = np.linalg.norm(eval_coord - ref_coord)

            # Find min dist
            if distance < min_distance:
                min_distance = distance

        # calc gamma index for this point
        gamma_index = min_distance / dta_thresh
        gamma_indexes.append(gamma_index)

    # Calculer la valeur moyenne du gamma index
    mean_gamma_index = np.mean(gamma_indexes)

    return mean_gamma_index


## TO FILL #######
plot_raw_resp = False
plot_strip_resp = False
fontsize_value = 10
CORRESPONDANCE_FILE = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
DATA_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\fx_formes_complexes\fx_circulaire\mesure\zData_150V_150ubeam_795mu_24p8v0_40collim17mmvitesse10.bin"
DCM_FILE = "RP.1.3.6.1.4.1.33868.20201021131037.169743"
NORMAL_VAL_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"

# strip pitch 0.2075 mm calculated at mask position
STRIP_PITCH = 0.2075
COUCH_SPEED = 10
THRESHOLD = 0.15

time_values, raw_strip_resp = read_data(DATA_FILE, CORRESPONDANCE_FILE)
couch_shift = np.array(time_values) * COUCH_SPEED
noize_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4)
normal_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 5)


# cut noize and normalize strip resp
noize_val, normal_val = get_and_format_noize_normal_val(
    NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4, time_values
)
strip_resp = (raw_strip_resp[18:153] - noize_val) / normal_val * 100
strip_resp = np.concatenate((np.zeros((18, len(time_values))), strip_resp))


# retrieves measurements whose resp > threshold and stores them in "points" array
points = np.empty((0, 3))
for istrip, strip in enumerate(strip_resp):
    for iresp, resp in enumerate(strip):
        if resp >= THRESHOLD:
            couch_mvt = time_values[iresp] * COUCH_SPEED
            dist_between_strips = istrip * STRIP_PITCH
            points = np.append(points, [(dist_between_strips, couch_mvt, resp)], axis=0)

# x and y = geometric coordinates of the points, v = response of the strip
x, y, v = points[:, 0], points[:, 1], points[:, 2]

# starts Y axis at 0
min_y = min(y)
y = y - min_y
points[:, 1] = points[:, 1] - min_y
couch_shift = couch_shift - min_y

# get center coord of theoratical circle beam shape
x_center_circ = get_x_coord_theo_circle(points)
y_center_circ = get_y_coord_theo_circle(x_center_circ, couch_shift, raw_strip_resp)

# center x axis on mid strip
points[:, 0] = points[:, 0] - x_center_circ
x_center_circ = 0

# create theoratical circle shape of 17 mm centered at calculated points
rad = 17 / 2
theta = np.linspace(0, 2 * np.pi, 100)
x_circ = x_center_circ + rad * np.cos(theta)
y_circ = y_center_circ + rad * np.sin(theta)

# gets miccrobeams x position
x_mb_coord = [x for [x, _, v] in points if v > 40]

# suppress every repeated x coord in x_mb_coord list
x_mb_coord = list(set(x_mb_coord))

# gets microbeam profile in Y
mb_v = []
mb_y = []
for x_val in x_mb_coord:
    # microbeam indexes
    indexes = np.where(x == x_val)[0]

    # corresponding v peak_values of microbeam considered
    mb_v_for_x_val = v[indexes]
    mb_y_for_x_val = y[indexes]

    # store v peak_values of microbeams in a list
    mb_v.append(mb_v_for_x_val)
    mb_y.append(mb_y_for_x_val)

# gets the half values of the microbeam FWHM
fwhm_values = [find_fwhm_values(mb) for mb in mb_v]

# gets the y coordinates of the half values of the microbeam FWHM
y_mb_coord = []

for index_mb, fwhm_tuple in enumerate(fwhm_values):
    # gets index of first half value in mb_v list
    index_y_half_val_1 = np.where(mb_v[index_mb] == fwhm_tuple[0])

    # gets y value corresponding to the index in mb_y list and adds it to y_mb_coord list
    y_mb_coord.append(mb_y[index_mb][index_y_half_val_1])

    # does the same for the second half value
    index_y_half_val_2 = np.where(mb_v[index_mb] == fwhm_tuple[1])
    y_mb_coord.append(mb_y[index_mb][index_y_half_val_2])

y_mb_coord = np.array(y_mb_coord)
# repeats every x coordinates of x_mb_coord twice (because for two y coord there is one
# x coord)
x_mb_coord_double = np.array([x for x in x_mb_coord for _ in range(2)])


# Plot theoratical and measured beam shapes
fig, ax = plt.subplots()
cmap = get_sub_cmap("YlGn", 0.2, 0.8)
norm = colors.Normalize(vmin=np.min(v), vmax=np.max(v))
ax.scatter(
    x_circ,
    y_circ,
)

# ax.plot(x_circ, y_circ, "k-", label="17 mm circle beam shape")
# im = ax.scatter(x, y, s=30, marker="s", cmap=cmap, vmin=v.min(), vmax=v.max(), c=v)
# cbar = fig.colorbar(im, label="Normalized response (%)")
# cbar.ax.tick_params(labelsize=fontsize_value)
# cbar.set_label("Normalized response (%)", fontsize=fontsize_value)
# ax.set_xlabel("Distance between strips at mask position (mm)", fontsize=fontsize_value)


# To get same scale on both axis
ticks = np.arange(-12, 30, 2)
ax.set_xticks(ticks)
ax.set_yticks(ticks)
ax.set_aspect("equal", adjustable="box")
ax.set_xlabel("Distance (mm)")
ax.set_ylabel("Distance (mm)")

# ax.scatter(x_circ, y_circ, s=30, marker="s", color="black")
ax.scatter(x_mb_coord_double, y_mb_coord, s=30, marker="s", color="red")
plt.show()

# Reorganize coordinates of theoratical shape and measured shape to have an array
# containing each coordinate couple [x, y]
theo_circ_coord = np.column_stack((x_circ, y_circ))
mes_circ_coord = np.column_stack((x_mb_coord_double, y_mb_coord))

# calc gamma index
dta_threshold = 0.5
coords1 = [(1, 1), (2, 2), (3, 3)]
coords2 = [(1.1, 1.1), (2.1, 2.1), (3.2, 3.3)]
gamma_index = calc_dta(coords1, coords2, dta_threshold)
# gamma_index = calc_dta(theo_circ_coord, mes_circ_coord, dta_threshold)
print(f"Gamma Index moyen: {gamma_index}")
