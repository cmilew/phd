import matplotlib.pyplot as plt
import matplotlib.colors as colors
import numpy as np
import math
import pydicom
import sys
from openpyxl import load_workbook
from scipy.interpolate import interp1d
from cmasher import get_sub_cmap
from scipy.signal import find_peaks, peak_widths


def read_data(data_file, correspondence_table_file):
    """Function to read data from measurement .bin file"""
    # .bin measurements file
    dt = np.dtype("uint16")
    f = open(data_file, "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)

    # correspondence of QDC number and strip number file
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_events = np.size(zdata) // 309

    # time conversion in seconds (integration time = 0.01s + 0.0005 s of dead time)
    time_values = [event * 0.0105 for event in range(nb_events)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, len(time_values)))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_events):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )

    return time_values, raw_strip_resp


def mask_hull(tuple_points):
    """Function to compute the convex hull of a set of points"""
    points = []
    for [x, y] in tuple_points:
        middle = math.ceil(len(points) / 2)
        if len(points) >= 2 and points[middle - 1][0] == x and points[middle][0] == x:
            points.pop(middle)
        points.insert(middle, (x, y))

    return points


def get_beam_contour_from_dcm(dcm_file):
    """Function to read dcm file containing beam shape (from TPS) and returning beam
    contours and isocenter position"""
    plan = pydicom.read_file(dcm_file, force=True)
    read_plan = pydicom.dcmread(dcm_file)
    dummy_X, dummy_Y, centroid = [], [], []
    beam_seq = plan[(0x300A, 0x00B0)].value
    beam_contours = []
    for i in range(len(beam_seq)):
        block_seq = beam_seq[i][(0x300A, 0x00F4)].value
        block_data = block_seq[0][(0x300A, 0x0106)].value
        dummy = []
        for entry in block_data:
            dummy.append(entry)
        dummy_array = np.array(dummy).astype(float)
        dummy_X.append(dummy_array[::2])
        dummy_X[i] = dummy_X[i].reshape(dummy_X[i].shape[0], 1)
        dummy_Y.append(dummy_array[1::2])
        dummy_Y[i] = dummy_Y[i].reshape(dummy_Y[i].shape[0], 1)
        dummy_X = dummy_X
        dummy_Y = dummy_Y
        centroid.append(
            (
                sum(dummy_X[i][:, 0]) / len(dummy_X[i][:, 0]),
                sum(dummy_Y[i][:, 0]) / len(dummy_Y[i][:, 0]),
            )
        )
        beam_contours.append(np.concatenate((dummy_X[i], dummy_Y[i]), axis=1))

    # Gets isocenter coordinates
    if hasattr(read_plan, "BeamSequence"):
        for beam in read_plan.BeamSequence:
            if hasattr(beam, "ControlPointSequence"):
                for cp in beam.ControlPointSequence:
                    if hasattr(cp, "IsocenterPosition"):
                        isocenter_pos = cp.IsocenterPosition  # [x, y, z]
    else:
        print("Pas d'information sur l'isocentre trouvée.")

    return beam_contours, isocenter_pos


def get_excel_data(excel_file, ws_name, start_l, start_col, n_l):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(excel_file)
    ws = wb_res[ws_name]
    excel_data = [ws.cell(row=start_l + i, column=start_col).value for i in range(n_l)]
    np.concatenate((np.zeros(18), excel_data))
    return np.array(excel_data)


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def get_and_format_noize_normal_val(
    excel_file, ws_name, start_l, start_col, time_values
):
    noize = get_excel_data(excel_file, ws_name, start_l, start_col, 135)
    normal_val = get_excel_data(excel_file, ws_name, start_l, start_col + 1, 135)
    noize = np.repeat(noize, len(time_values)).reshape(-1, len(time_values))
    noize_strips = np.concatenate((np.zeros((18, len(time_values))), noize))
    normal_val = np.repeat(normal_val, len(time_values)).reshape(-1, len(time_values))
    normal_val_strips = np.concatenate((np.zeros((18, len(time_values))), normal_val))

    return noize_strips, normal_val_strips


def plot_resp(time, resp, title, y_label):
    for strip_num in range(18, 153):
        plt.plot(time, resp[strip_num, :], label=f"Strip {strip_num}")
    plt.xlabel("Time (s)")
    plt.ylabel(y_label)
    plt.legend(loc="best")
    plt.title(title, fontsize=18, fontstyle="oblique")
    plt.show()


def get_central_absc_val_fwhm(x, y):
    """Function to return the abscissa of the FWHM central value of a profile"""

    max_index = y.argmax()
    fwhm = peak_widths(y, [max_index], rel_height=0.5)
    fwhm_left = np.interp(fwhm[2][0], np.arange(len(x)), x)
    fwhm_right = np.interp(fwhm[3][0], np.arange(len(x)), x)

    return (fwhm_right - fwhm_left) / 2 + fwhm_left


def resample_points_at_given_x(points, x_target):
    """Function resampling a given set of coordinates (x,y) by linear interpolation at given x values"""
    x_target = np.array(x_target)
    x = points[:, 0]
    y = points[:, 1]
    f = interp1d(x, y, kind="linear", fill_value="extrapolate")
    y_target = f(x_target)

    return np.column_stack((x_target, y_target))


def cut_cont_in_between_thresh(contours, thres_low, thresh_high):
    """Function cutting contours in two part : one top and one bottom at a given
    threshold"""

    maks = (contours[:, 1] > thres_low) & (contours[:, 1] < thresh_high)
    contours_cut = contours[maks]

    return contours_cut


def cut_beam_contours_at_given_x(beam_contours, x_cut):
    """Function returning beam_contours cut at x_cut"""

    # retrieve coordinates of point located before x_cut
    beam_contours_cut = np.column_stack(
        (
            beam_contours[beam_contours[:, 0] < x_cut][:, 0],
            beam_contours[beam_contours[:, 0] < x_cut][:, 1],
        )
    )

    # cut beam contours in 2 : one top and one bottom part to interpolate separately
    # edges of cut shape, thresh y = -6
    beam_contours_top = np.column_stack(
        (
            beam_contours[beam_contours[:, 1] > -6][:, 0],
            beam_contours[beam_contours[:, 1] > -6][:, 1],
        )
    )

    beam_contours_bott = np.column_stack(
        (
            beam_contours[beam_contours[:, 1] < -6][:, 0],
            beam_contours[beam_contours[:, 1] < -6][:, 1],
        )
    )

    # get top edge contour
    top_interpolation = interp1d(
        beam_contours_top[:, 0],
        beam_contours_top[:, 1],
        kind="linear",
        fill_value="extrapolate",
    )
    y_cut_top = top_interpolation(x_cut)

    # adds top edge contour to arrays of the cut shape coordinates at right place to
    # keep order of points
    beam_contours_cut = np.insert(beam_contours_cut, 14, [x_cut, y_cut_top], axis=0)

    # get bottom edge contour
    bott_interpolation = interp1d(
        beam_contours_bott[:, 0],
        beam_contours_bott[:, 1],
        kind="linear",
        fill_value="extrapolate",
    )
    y_cut_bott = bott_interpolation(x_cut)

    # adds top edge contour to arrays of the cut shape coordinates at right place
    beam_contours_cut = np.insert(beam_contours_cut, 14, [x_cut, y_cut_bott], axis=0)

    return beam_contours_cut


def reg_beam_cont_to_a_given_point(beam_cont, ref_point, point_to_reg):
    """Function calculating the x_shift and y_shift necessery to match point_to_reg to
    ref_point and applying the same shifts to every coordinate of beam_cont"""

    x_shift = ref_point[0] - point_to_reg[0]
    y_shift = ref_point[1] - point_to_reg[1]

    x_reg = beam_cont[:, 0] + x_shift
    y_reg = beam_cont[:, 1] + y_shift

    return [point_to_reg[0] + x_shift, point_to_reg[1] + y_shift], np.column_stack(
        (x_reg, y_reg)
    )


def find_fwhm_values(peak_values):
    """Function returning the left and right half-maximum peak_values of a peak"""
    # Convert to numpy array for easier manipulation
    peak_values = np.array(peak_values)

    # Find the index of the peak
    peak_index = np.argmax(peak_values)
    peak_value = peak_values[peak_index]

    # Calculate the half maximum
    half_max = peak_value / 2.0

    # Find the left half-maximum point
    left_index = np.where(peak_values[:peak_index] <= half_max)[0][-1]
    right_index = np.where(peak_values[peak_index:] <= half_max)[0][0] + peak_index

    left_hm = peak_values[left_index]
    right_hm = peak_values[right_index]

    return left_hm, right_hm


def get_meas_cont(points):
    # gets miccrobeams x position
    x_mb_coord = [x for [x, _, v] in points if v > 40]

    # suppress every repeated x coord in x_mb_coord list
    x_mb_coord = list(set(x_mb_coord))

    # gets microbeam profile in Y
    mb_v = []
    mb_y = []
    for x_val in x_mb_coord:
        # corresponding y and v peak_values of microbeam considered
        mb_v_for_x_val = points[points[:, 0] == x_val][:, 2]
        mb_y_for_x_val = points[points[:, 0] == x_val][:, 1]

        # store v and peak_values of microbeams in a list
        mb_v.append(mb_v_for_x_val)
        mb_y.append(mb_y_for_x_val)

    # gets the half values of the microbeam FWHM
    fwhm_values = [find_fwhm_values(mb) for mb in mb_v]

    # gets the y coordinates of the half values of the microbeam FWHM
    y_mb_coord = []

    for index_mb, fwhm_tuple in enumerate(fwhm_values):
        # gets index of first half value in mb_v list
        index_y_half_val_1 = np.where(mb_v[index_mb] == fwhm_tuple[0])

        # gets y value corresponding to the index in mb_y list and adds it to y_mb_coord list
        y_mb_coord.append(mb_y[index_mb][index_y_half_val_1])

        # does the same for the second half value
        index_y_half_val_2 = np.where(mb_v[index_mb] == fwhm_tuple[1])
        y_mb_coord.append(mb_y[index_mb][index_y_half_val_2])

    y_mb_coord = np.array(y_mb_coord)
    # repeats every x coordinates of x_mb_coord twice (because for two y coord there is one
    # x coord)
    x_mb_coord_double = np.array([x for x in x_mb_coord for _ in range(2)])

    return np.column_stack((x_mb_coord_double, y_mb_coord))


# TO FILL #############
plot_raw_resp = False
plot_strip_resp = False
fontsize_value = 10
# strip measuring the microbeam on the furthest right of the mask
extreme_right_strip = 146

CORRESPONDANCE_FILE = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
DATA_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\fx_formes_complexes\fx_cochons\mesures\zData_150V_150ubeam_795mu_24p8v0_40_collimCochon_vitesse10.bin"
DCM_FILE = r"C:\Users\milewski\Desktop\these\phd\codes_python\fx_formes_complexes\RP.1.3.6.1.4.1.33868.20210210162030.730598"
NORMAL_VAL_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"

COUCH_SPEED = 10
THRESHOLD = 0.15

# strip exact positioning
strip_pos_file = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\microbeam_scan_analysis\strip_exact_positioning.xlsx"
strip_pos = get_excel_data(strip_pos_file, "strip_pos", 5, 7, 153)


# strip positionning at mask
pitch_strip = 0.2325
pitch_at_mask = 0.2075
dim_fact = pitch_at_mask / pitch_strip
strip_pos_at_mask = strip_pos * dim_fact
dic_strip_pos = {i: strip_pos_at_mask[i] for i in range(153)}
central_strip = 73 - 1

time_values, raw_strip_resp = read_data(DATA_FILE, CORRESPONDANCE_FILE)
couch_shift = np.array(time_values) * COUCH_SPEED
noize_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4, 135)
normal_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 5, 135)


# raw strips response plot
if plot_raw_resp:
    plot_resp(
        time_values,
        raw_strip_resp,
        "Raw responses pig beam ESRF",
        "QDC response (AU)",
    )


# cut noize and normalize strip resp
noize_val, normal_val = get_and_format_noize_normal_val(
    NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4, time_values
)
strip_resp = (raw_strip_resp - noize_val) / normal_val * 100

# strips response plot
if plot_strip_resp:
    plot_resp(
        time_values,
        strip_resp,
        "Normalized strip responses pig beam ESRF",
        "Normalized resp (%)",
    )


# Compute points
points = np.empty((0, 3))
for istrip, strip in enumerate(strip_resp):
    for iresp, resp in enumerate(strip):
        if resp >= THRESHOLD:
            couch_mvt = time_values[iresp] * COUCH_SPEED
            # dist_between_strips = istrip * STRIP_PITCH
            points = np.append(
                points, [(dic_strip_pos[istrip], couch_mvt, resp)], axis=0
            )

x, y, v = points[:, 0], points[:, 1], points[:, 2]

# to start Y axis at 0
min_y = min(y)
y = [y[i] - min_y for i in range(len(y))]
points[:, 1] = points[:, 1] - min_y
couch_shift = couch_shift - min_y

# beam contours from dcm
beam_cont_dcm, iso_pos = get_beam_contour_from_dcm(DCM_FILE)
beam_contours = beam_cont_dcm[1][:-1]

# add first point of beam contours at the end to complete the shape
beam_contours = np.vstack([beam_contours, beam_contours[0]])

# reverse beam contours shape to retrieve same orientation as strip measurements
beam_contours[:, 1] = -beam_contours[:, 1]
beam_contours[:, 0] = -beam_contours[:, 0]
iso_pos[1] = -iso_pos[1]
iso_pos[0] = -iso_pos[0]


# first and last strip measuring microbeam signal = 18 and 146
beam_width_seen = abs(dic_strip_pos[146] - dic_strip_pos[18])

# finding at which x coord to cut beam contours to keep only the part of the beam shape
# seen by detector
beam_width = np.max(beam_contours[:, 0]) - np.min(beam_contours[:, 0])
width_to_cut = beam_width - beam_width_seen
x_cut = np.max(beam_contours[:, 0]) - width_to_cut

beam_contours_cut = cut_beam_contours_at_given_x(beam_contours, x_cut)

# cut beam contours in half
# beam_cont_top, beam_cont_bott = cut_cont_in_half(beam_contours_cut, -6)

# # get measured contours
# meas_cont = get_meas_cont(points)
# meas_barycenter = [np.mean(meas_cont[:, 0]), np.mean(meas_cont[:, 1])]

# # retrieves x to which interpolate beam contours
# x_target_resample = np.unique(meas_cont[:, 0])

# # resample beam contours at x_target_resample
# beam_cont_top_resamp = resample_points_at_given_x(beam_cont_top, x_target_resample)
# beam_cont_bott_resamp = resample_points_at_given_x(beam_cont_bott, x_target_resample)

# plt.scatter(beam_cont_top_resamp[:, 0], beam_cont_top_resamp[:, 1], c="blue")
# plt.scatter(beam_cont_bott_resamp[:, 0], beam_cont_bott_resamp[:, 1], c="red")
plt.scatter(beam_contours_cut[:, 0], beam_contours_cut[:, 1], c="green")
plt.show()


sys.exit()


# calc barycenter of beam contours cut
tps_barycenter = [np.mean(beam_contours_cut[:, 0]), np.mean(beam_contours_cut[:, 1])]


# try registering barycenter of beam shape TPS to measured barycenter
new_iso, beam_cont_reg_iso = reg_beam_cont_to_a_given_point(
    beam_contours_cut, meas_barycenter, tps_barycenter
)


fontsize_value = 15
fig, ax = plt.subplots()
ax.scatter(beam_contours_cut[:, 0], beam_contours_cut[:, 1], c="blue")
# ax.scatter(iso_pos[0], iso_pos[1], c="red", label="Isocenter")
ax.scatter(tps_barycenter[0], tps_barycenter[1], c="black", label="Barycenter")
plt.show()

fig, ax = plt.subplots()
# ax.plot(beam_cont_reg_iso[:, 0], beam_cont_reg_iso[:, 1], "k-")
ax.scatter(meas_cont[:, 0], meas_cont[:, 1], color="red", label="isocenter")
ax.scatter(
    meas_barycenter[0],
    meas_barycenter[1],
    color="black",
    label="barycentre forme mesurée",
)
plt.show()
sys.exit()

cmap = get_sub_cmap("YlGn", 0.2, 0.8)
norm = colors.Normalize(vmin=np.min(v), vmax=np.max(v))
im = ax.scatter(x, y, s=5, cmap=cmap, vmin=v.min(), vmax=v.max(), c=v)
cbar = fig.colorbar(im, label="Normalized response (%)")
cbar.ax.tick_params(labelsize=fontsize_value)
cbar.set_label("Normalized response (%)", fontsize=fontsize_value)
ax.set_xlabel("Distance between strips at mask position (mm)", fontsize=fontsize_value)
ax.legend(fontsize=fontsize_value)
# plt.scatter(dic_strip_pos[18], 5)
# plt.scatter(dic_strip_pos[146], 15)
# plt.scatter(dic_strip_pos[85], y_center_mid_strip, c="red")


# # To get same scale on both axis
# ticks = np.arange(0, 38, 2)
# ax.set_xticks(ticks)
# ax.set_yticks(ticks)
# ax.set_aspect("equal", adjustable="box")
# ax.set_ylabel("Couch shift (mm)", fontsize=fontsize_value)
# ax.set_ylim(-2, 24)
# ax.set_xlim(0, 34)
plt.scatter(new_iso[0], new_iso[1], c="red", label="Isocenter")
plt.show()
