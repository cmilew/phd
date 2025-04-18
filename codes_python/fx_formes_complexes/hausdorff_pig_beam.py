import matplotlib.pyplot as plt
import matplotlib.colors as colors
import numpy as np
import pydicom
import sys
from openpyxl import load_workbook
from scipy.interpolate import interp1d
from cmasher import get_sub_cmap
from scipy.spatial import cKDTree
from scipy.spatial.distance import directed_hausdorff


def read_data(data_file, correspondence_table_file):
    """Function to read data from measurement .bin file"""
    # .bin measurements file
    dt = np.dtype("uint16")
    f = open(data_file, "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)

    # correspondence of QDC number and strip number file
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_events = np.size(zdata) // 309

    # time conversion in seconds (integration time = 0.01s + 0.0005 s of dead time)
    time_values = [event * 0.0105 for event in range(nb_events)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, len(time_values)))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_events):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )

    return time_values, raw_strip_resp


def get_beam_contour_from_dcm(dcm_file):
    """Function to read dcm file containing beam shape (from TPS) and returning beam
    contours and isocenter position"""
    plan = pydicom.read_file(dcm_file, force=True)
    read_plan = pydicom.dcmread(dcm_file)
    dummy_X, dummy_Y, centroid = [], [], []
    beam_seq = plan[(0x300A, 0x00B0)].value
    beam_contours = []
    for i in range(len(beam_seq)):
        block_seq = beam_seq[i][(0x300A, 0x00F4)].value
        block_data = block_seq[0][(0x300A, 0x0106)].value
        dummy = []
        for entry in block_data:
            dummy.append(entry)
        dummy_array = np.array(dummy).astype(float)
        dummy_X.append(dummy_array[::2])
        dummy_X[i] = dummy_X[i].reshape(dummy_X[i].shape[0], 1)
        dummy_Y.append(dummy_array[1::2])
        dummy_Y[i] = dummy_Y[i].reshape(dummy_Y[i].shape[0], 1)
        dummy_X = dummy_X
        dummy_Y = dummy_Y
        centroid.append(
            (
                sum(dummy_X[i][:, 0]) / len(dummy_X[i][:, 0]),
                sum(dummy_Y[i][:, 0]) / len(dummy_Y[i][:, 0]),
            )
        )
        beam_contours.append(np.concatenate((dummy_X[i], dummy_Y[i]), axis=1))

    # Gets isocenter coordinates
    if hasattr(read_plan, "BeamSequence"):
        for beam in read_plan.BeamSequence:
            if hasattr(beam, "ControlPointSequence"):
                for cp in beam.ControlPointSequence:
                    if hasattr(cp, "IsocenterPosition"):
                        isocenter_pos = cp.IsocenterPosition  # [x, y, z]
    else:
        print("Pas d'information sur l'isocentre trouvée.")

    beam_cont = beam_contours[1][:-1]

    # add first point of beam contours at the end to complete the shape
    beam_cont = np.vstack([beam_cont, beam_cont[0]])

    return beam_cont, isocenter_pos


def get_excel_data(excel_file, ws_name, start_l, start_col, n_l):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(excel_file)
    ws = wb_res[ws_name]
    excel_data = [ws.cell(row=start_l + i, column=start_col).value for i in range(n_l)]
    np.concatenate((np.zeros(18), excel_data))
    return np.array(excel_data)


def get_and_format_noize_normal_val(
    excel_file, ws_name, start_l, start_col, time_values
):
    noize = get_excel_data(excel_file, ws_name, start_l, start_col, 135)
    normal_val = get_excel_data(excel_file, ws_name, start_l, start_col + 1, 135)
    noize = np.repeat(noize, len(time_values)).reshape(-1, len(time_values))
    noize_strips = np.concatenate((np.zeros((18, len(time_values))), noize))
    normal_val = np.repeat(normal_val, len(time_values)).reshape(-1, len(time_values))
    normal_val_strips = np.concatenate((np.zeros((18, len(time_values))), normal_val))

    return noize_strips, normal_val_strips


def cut_cont_in_half(contours, thresh):
    """Function cutting contours in two part : one top and one bottom at a given
    threshold"""

    maks = contours[:, 1] > thresh
    contours_top = contours[maks]
    contours_bott = contours[~maks]

    # delete duplicate in contours top
    contours_top = np.unique(contours_top, axis=0)

    return contours_top, contours_bott


def reverse_beam_cont(beam_contours, iso_pos):
    """Function reversing beam contours shape to retrieve same orientation as strip
    measurements"""
    beam_contours[:, 1] = -beam_contours[:, 1]
    beam_contours[:, 0] = -beam_contours[:, 0]
    iso_pos[1] = -iso_pos[1]
    iso_pos[0] = -iso_pos[0]

    return beam_contours, iso_pos


def cut_beam_contours_at_given_x(beam_contours, x_cut):
    """Function returning beam_contours cut at x_cut"""

    # retrieve coordinates of point located before x_cut
    beam_contours_cut = np.column_stack(
        (
            beam_contours[beam_contours[:, 0] < x_cut][:, 0],
            beam_contours[beam_contours[:, 0] < x_cut][:, 1],
        )
    )

    # cut beam contours in 2 : one top and one bottom part to interpolate separately
    # edges of cut shape, thresh y = -6
    beam_cont_cut_top, beam_cont_cut_bott = cut_cont_in_half(beam_contours_cut, 6)

    # get top edge contour
    top_interpolation = interp1d(
        beam_cont_cut_top[:, 0],
        beam_cont_cut_top[:, 1],
        kind="linear",
        fill_value="extrapolate",
    )
    y_cut_top = top_interpolation(x_cut)

    # adds top edge contour to arrays of the cut shape coordinates at right place to
    # keep order of points
    beam_contours_cut = np.insert(beam_contours_cut, 14, [x_cut, y_cut_top], axis=0)

    # get bottom edge contour
    bott_interpolation = interp1d(
        beam_cont_cut_bott[:, 0],
        beam_cont_cut_bott[:, 1],
        kind="linear",
        fill_value="extrapolate",
    )
    y_cut_bott = bott_interpolation(x_cut)

    # adds top edge contour to arrays of the cut shape coordinates at right place
    beam_contours_cut = np.insert(beam_contours_cut, 14, [x_cut, y_cut_bott], axis=0)

    return beam_contours_cut


def reg_beam_cont_to_a_given_point(beam_cont, ref_point, point_to_reg):
    """Function calculating the x_shift and y_shift necessery to match point_to_reg to
    ref_point and applying the same shifts to every coordinate of beam_cont"""

    x_shift = ref_point[0] - point_to_reg[0]
    y_shift = ref_point[1] - point_to_reg[1]

    x_reg = beam_cont[:, 0] + x_shift
    y_reg = beam_cont[:, 1] + y_shift

    return [point_to_reg[0] + x_shift, point_to_reg[1] + y_shift], np.column_stack(
        (x_reg, y_reg)
    )


def find_fwhm_values(peak_values):
    """Function returning the left and right half-maximum peak_values of a peak"""
    # Convert to numpy array for easier manipulation
    peak_values = np.array(peak_values)

    # Find the index of the peak
    peak_index = np.argmax(peak_values)
    peak_value = peak_values[peak_index]

    # Calculate the half maximum
    half_max = peak_value / 2.0

    # Find the left half-maximum point
    left_index = np.where(peak_values[:peak_index] <= half_max)[0][-1]
    right_index = np.where(peak_values[peak_index:] <= half_max)[0][0] + peak_index

    left_hm = peak_values[left_index]
    right_hm = peak_values[right_index]

    return left_hm, right_hm


def get_meas_cont(points):
    """Function to get the measured contours by the detector. Measured contours
    correspond to the FWHM values of each microbeam y profiles"""

    # gets miccrobeams x position
    x_mb_coord = [x for [x, _, v] in points if v > 40]

    # suppress every repeated x coord in x_mb_coord list
    x_mb_coord = list(set(x_mb_coord))

    # gets microbeam profile in Y
    mb_v = []
    mb_y = []
    for x_val in x_mb_coord:
        # corresponding y and v peak_values of microbeam considered
        mb_v_for_x_val = points[points[:, 0] == x_val][:, 2]
        mb_y_for_x_val = points[points[:, 0] == x_val][:, 1]

        # store v and peak_values of microbeams in a list
        mb_v.append(mb_v_for_x_val)
        mb_y.append(mb_y_for_x_val)

    # gets the half values of the microbeam FWHM
    fwhm_values = [find_fwhm_values(mb) for mb in mb_v]

    # gets the y coordinates of the half values of the microbeam FWHM
    y_mb_coord = []

    for index_mb, fwhm_tuple in enumerate(fwhm_values):
        # gets index of first half value in mb_v list
        index_y_half_val_1 = np.where(mb_v[index_mb] == fwhm_tuple[0])

        # gets y value corresponding to the index in mb_y list and adds it to y_mb_coord list
        y_mb_coord.append(mb_y[index_mb][index_y_half_val_1])

        # does the same for the second half value
        index_y_half_val_2 = np.where(mb_v[index_mb] == fwhm_tuple[1])
        y_mb_coord.append(mb_y[index_mb][index_y_half_val_2])

    y_mb_coord = np.array(y_mb_coord)
    # repeats every x coordinates of x_mb_coord twice (because for two y coord there is one
    # x coord)
    x_mb_coord_double = np.array([x for x in x_mb_coord for _ in range(2)])

    return np.column_stack((x_mb_coord_double, y_mb_coord))


def separate_cont_in_two_parts(cont):
    """Function separating contours in two parts : one top and one bottom but taking
    into account that it's a complex shape and the threshold between top and bott has
    to be adjusted depending of the x coordinate of the contours"""

    # mask to separate top from bott contours
    # top cont = all coord for which when x < 10 y > 12 and when x > 10, y > 6
    mask = ((cont[:, 0] < 10) & (cont[:, 1] > 12)) | (
        (cont[:, 0] > 10) & (cont[:, 1] > 6)
    )
    top_cont = cont[mask]
    # bott contour = every other coord
    bott_cont = cont[~mask]

    return top_cont, bott_cont


def find_y_target_iso_reg(mes_cont, x_target_iso, d_cont_top_iso, d_cont_bott_iso):
    """Function finding the target y pos where iso should be move to register TPS cont
    to measured contours. It correspond to the mean value of the distance between the
    top and bott contours and the central strip y profile FWHM values"""

    # finds FWHM values of central strip y profile
    y_at_center_strip = mes_cont[mes_cont[:, 0] == x_target_iso, 1]

    # find target y positions for isocenter registration = should be at the same distance
    # from top and bott contours as the central strip y profile as the iso is from its TPS
    # contours (but not exactly so mean value is taken)
    y_target_iso_1 = y_at_center_strip[0] + d_cont_bott_iso
    y_target_iso_2 = y_at_center_strip[1] - d_cont_top_iso
    y_target_iso = (y_target_iso_1 + y_target_iso_2) / 2

    return y_target_iso


def resample_at_given_x(cont, x_resample):
    """Function resampling a given set of coordinates (x,y) by linear interpolation at
    given x values"""

    # makes sure coord are sorted by x to prevent interpolation error
    cont = cont[np.argsort(cont[:, 0])]

    # linear interpol y values for new x values
    y_resample = np.interp(x_resample, cont[:, 0], cont[:, 1])

    cont_resample = np.column_stack((x_resample, y_resample))

    # check if they are duplicates in x coord => means there are vertical line in beam
    # shape
    unique_x, counts = np.unique(cont[:, 0], return_counts=True)
    x_duplicates = unique_x[counts > 1]

    for x_dup in x_duplicates:
        # retrieves y values of x duplicates
        y_duplicates = cont[np.where(cont[:, 0] == x_dup)][:, 1]
        assert len(y_duplicates) == 2, "should only be 2 y coord for every x duplicates"

        # resample vertical line
        x_resample_vert_line = np.repeat(x_dup, 100)
        y_resample_vert_line = np.linspace(max(y_duplicates), min(y_duplicates), 100)

        # arrange new coordinates
        cont_resample_vert_line = np.column_stack(
            (x_resample_vert_line, y_resample_vert_line)
        )

        # concatenate with resampled contours
        cont_resample = np.vstack((cont_resample, cont_resample_vert_line))

    # Concatenate contours with all resampled contours
    all_cont = np.vstack((cont, cont_resample))

    # makes sure coord are sorted by x to keep coherence
    all_cont = all_cont[np.argsort(all_cont[:, 0])]
    return all_cont


def find_nearest_points(array_1, array_2):
    """
    Trouve les points de array_1 les plus proches des points de array_2.

    :param array_1: np.array de forme (N, 2) contenant les coordonnées [x, y]
    :param array_2: np.array de forme (M, 2) contenant les coordonnées [x, y]
    :return: np.array des points de array_1 correspondant aux plus proches voisins de array_2
    """
    tree = cKDTree(array_1)  # Construire un arbre KD pour une recherche efficace
    _, indices = tree.query(array_2)  # Trouver les indices des plus proches voisins

    return array_1[indices]  # Retourner les points correspondants


def hausdorff_distance(set1, set2):
    """
    Calculate the Hausdorff distance between two sets of points.

    Parameters:
    set1 (numpy.ndarray): First set of points, shape (n_points1, 2)
    set2 (numpy.ndarray): Second set of points, shape (n_points2, 2)

    Returns:
    float: Hausdorff distance between set1 and set2
    """
    # Compute the directed Hausdorff distances
    d1 = directed_hausdorff(set1, set2)[0]
    d2 = directed_hausdorff(set2, set1)[0]

    # The Hausdorff distance is the maximum of the two directed distances
    return max(d1, d2)


# TO FILL #############
plot_raw_resp = False
plot_strip_resp = False
fontsize_value = 20

CORRESPONDANCE_FILE = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
DATA_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\fx_formes_complexes\fx_cochons\mesures\zData_150V_150ubeam_795mu_24p8v0_40_collimCochon_vitesse10.bin"
DCM_FILE = r"C:\Users\milewski\Desktop\these\phd\codes_python\fx_formes_complexes\RP.1.3.6.1.4.1.33868.20210210162030.730598"
NORMAL_VAL_FILE = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"

COUCH_SPEED = 10
THRESHOLD = 0.15

# strip exact positioning
strip_pos_file = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\microbeam_scan_analysis\strip_exact_positioning.xlsx"
strip_pos = get_excel_data(strip_pos_file, "strip_pos", 5, 7, 153)


# strip positionning at mask
pitch_strip = 0.2325
pitch_at_mask = 0.2075
dim_fact = pitch_at_mask / pitch_strip
strip_pos_at_mask = strip_pos * dim_fact
dic_strip_pos = {i: strip_pos_at_mask[i] for i in range(153)}
central_strip = 73 - 1

time_values, raw_strip_resp = read_data(DATA_FILE, CORRESPONDANCE_FILE)
couch_shift = np.array(time_values) * COUCH_SPEED
noize_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4, 135)
normal_val = get_excel_data(NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 5, 135)


# cut noize and normalize strip resp
noize_val, normal_val = get_and_format_noize_normal_val(
    NORMAL_VAL_FILE, "mb_scan_ESRF", 4, 4, time_values
)
strip_resp = (raw_strip_resp - noize_val) / normal_val * 100

# Compute points
points = np.empty((0, 3))
for istrip, strip in enumerate(strip_resp):
    for iresp, resp in enumerate(strip):
        if resp >= THRESHOLD:
            couch_mvt = time_values[iresp] * COUCH_SPEED
            # dist_between_strips = istrip * STRIP_PITCH
            points = np.append(
                points, [(dic_strip_pos[istrip], couch_mvt, resp)], axis=0
            )

x, y, v = points[:, 0], points[:, 1], points[:, 2]

# to start Y axis at 0
min_y = min(y)
y = [y[i] - min_y for i in range(len(y))]
points[:, 1] = points[:, 1] - min_y
couch_shift = couch_shift - min_y

# beam contours from dcm
beam_contours, iso_pos = get_beam_contour_from_dcm(DCM_FILE)

# reverse beam contours shape to retrieve same orientation as strip measurements
beam_contours, iso_pos = reverse_beam_cont(beam_contours, iso_pos)

# cut beam contours in half
cont_top, cont_bott = cut_cont_in_half(beam_contours, 0)
cont_top = np.unique(cont_top, axis=0)

# get y coordinates of top and bott contours for x of iso pos
y_mid_strip_top = np.interp(iso_pos[0], cont_top[:, 0], cont_top[:, 1])
y_mid_strip_bott = np.interp(iso_pos[0], cont_bott[:, 0], cont_bott[:, 1])

# get distance between top and bott contours and iso pos to know where iso is located
# relatively to these the contour above and below it
d_cont_top_iso = abs(y_mid_strip_top - iso_pos[1])
d_cont_bott_iso = abs(y_mid_strip_bott - iso_pos[1])

# get meas cont (=FWHM values of each y profile of strip response)
meas_cont = get_meas_cont(points)

# find where iso of TPS cont should be positioned to be register to measured contours
# x target for iso = x of central strip
x_target_iso = dic_strip_pos[central_strip]

# calc the y target position for iso on the 2 strips neighboring the central strip
# (which are in the beam) and interpolate in between
x_target_right_strip = dic_strip_pos[central_strip - 1]
y_target_right_strip = find_y_target_iso_reg(
    meas_cont, x_target_right_strip, d_cont_top_iso, d_cont_bott_iso
)
x_target_left_strip = dic_strip_pos[central_strip + 1]
y_target_left_strip = find_y_target_iso_reg(
    meas_cont, x_target_left_strip, d_cont_top_iso, d_cont_bott_iso
)

# interpolate in between
y_target_iso = np.interp(
    x_target_iso,
    [x_target_left_strip, x_target_right_strip],
    [y_target_left_strip, y_target_right_strip],
)

# register TPS isocenter to center of middle strip
new_iso, beam_cont_reg = reg_beam_cont_to_a_given_point(
    beam_contours, [x_target_iso, y_target_iso], [iso_pos[0], iso_pos[1]]
)

# x coord swhere to cut TPS beam shape to compare shapes with Hausdorff distance
x_cut = dic_strip_pos[18]
beam_cont_cut = cut_beam_contours_at_given_x(beam_cont_reg, x_cut)

# suppress last points because it's the same as the first one
beam_cont_cut = np.delete(beam_cont_cut, -1, axis=0)
beam_cont_top, beam_cont_bott = separate_cont_in_two_parts(beam_cont_cut)

# exchange first and second coordinates of beam_cont_top so that y are in increasing
# order
beam_cont_top = beam_cont_top[np.argsort(beam_cont_top[:, 0])]
aux = beam_cont_top[0].copy()
beam_cont_top[0] = beam_cont_top[1]
beam_cont_top[1] = aux

# add last point of beam_cont_bott to beam_cont_top for resampling interpolation
beam_cont_top = np.vstack([beam_cont_top, beam_cont_bott[-1]])

# add first point of beam_cont_top to beam_cont_bott for resampling interpolation
beam_cont_bott = np.vstack([beam_cont_top[0], beam_cont_bott])

# resample beam contours to have more points of comparison to measured contours
x_resample_top = np.linspace(min(beam_cont_top[:, 0]), max(beam_cont_top[:, 0]), 10000)
beam_cont_top_resample = resample_at_given_x(beam_cont_top, x_resample_top)
x_resample_bott = np.linspace(
    min(beam_cont_bott[:, 0]), max(beam_cont_bott[:, 0]), 10000
)
beam_cont_bott_resample = resample_at_given_x(beam_cont_bott, x_resample_bott)

# concatenate top and bott resampled contours
beam_cont_resample = np.vstack([beam_cont_top_resample, beam_cont_bott_resample])

# find closest coordinates of resampled beam contours to measured contours (to calc
# Hausdorff distance with same sampling)
beam_cont_same_sample = find_nearest_points(beam_cont_resample, meas_cont)

# plot measured contours and TPS beam contours
fontsize_value = 15
fig, ax = plt.subplots()
cmap = get_sub_cmap("YlGn", 0.2, 0.8)
norm = colors.Normalize(vmin=np.min(v), vmax=np.max(v))
im = ax.scatter(x, y, s=5, cmap=cmap, vmin=v.min(), vmax=v.max(), c=v)
cbar = fig.colorbar(im, label="Normalized response (%)")
cbar.ax.tick_params(labelsize=fontsize_value)
cbar.set_label("Normalized response (%)", fontsize=fontsize_value)
ax.set_xlabel("Distance between strips at mask position (mm)", fontsize=fontsize_value)
ax.scatter(
    meas_cont[:, 0], meas_cont[:, 1], color="r", marker="s", label="measured contours"
)
ax.scatter(
    beam_cont_same_sample[:, 0],
    beam_cont_same_sample[:, 1],
    label="theoratical contours",
)
ax.legend(fontsize=fontsize_value)

# To get same scale on both axis
x_ticks = np.arange(-16, 16, 2)
y_ticks = np.arange(0, 24, 2)
ax.set_xticks(x_ticks)
ax.set_yticks(y_ticks)
ax.set_aspect("equal", adjustable="box")
ax.set_ylabel("Couch shift (mm)", fontsize=fontsize_value)
ax.tick_params(axis="both", which="major", labelsize=fontsize_value)
plt.show()

# Calc Hausdorff distance
distance = hausdorff_distance(beam_cont_same_sample, meas_cont)
print(f"Hausdorff distance: {distance} mm")
