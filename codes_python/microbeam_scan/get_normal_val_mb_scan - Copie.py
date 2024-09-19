import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.signal import find_peaks, peak_widths
import sys


#### TO FILL ####
fontsize_value = 25
fill_excel_data = False
plot_raw_resp = False
plot_strip_resp = False


def read_data(data_file, correspondence_table_file):
    """Function to read data from measurement .bin file"""
    # .bin measurements file
    dt = np.dtype("uint16")
    f = open(data_file, "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)

    # correspondence of QDC number and strip number file
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_events = np.size(zdata) // 309

    # time conversion in seconds (integration time = 0.01s + 0.0005 s of dead time)
    time_values = [event * 0.0105 for event in range(nb_events)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, len(time_values)))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_events):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )

    return time_values, raw_strip_resp


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l + i, column=start_col, value=data_to_fill[i])
    wb_res.save(excel_path)


def calc_normal_val(time, resp):
    """Function calculating the normalization value of each strip which corresponds to
    the center of the peak response"""

    normal_val = np.zeros(153)

    for strip in range(18, 153):
        # find peak in strip resp
        peak_index, _ = find_peaks(resp[strip], height=2500)
        assert len(peak_index) == 1, "More than one peak found for strip {}".format(
            strip
        )

        # find FWHM
        fwhm = peak_widths(resp[strip], peak_index, rel_height=0.5)
        fwhm_left = np.interp(fwhm[2][0], np.arange(len(time)), time)
        fwhm_right = np.interp(fwhm[3][0], np.arange(len(time)), time)

        # calc coord of mid FWHM = center value of peak
        x_mid_peak = (fwhm_left + fwhm_right) / 2
        y_mid_peak = np.interp(x_mid_peak, time_values, strip_resp[strip])

        normal_val[strip] = y_mid_peak
    return normal_val


# read measurements data
mes_file = r"C:\Users\milewski\Desktop/these/mesures/caracterisation_detecteur_8-06-2023/more_homogeneous_zData_150V_1ubeam_24p8v0_scan6.bin"
correspondence_table_file = r"C:\Users\milewski\Desktop/these/mesures/analyse_data/codes_python/150_voies/add_piste.txt"
res_excel_file = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"

time_values, raw_strip_resp = read_data(mes_file, correspondence_table_file)
fontsize_value = 10

# raw strips response plot
if plot_raw_resp:
    for strip_num in range(18, 153):
        plt.plot(time_values, raw_strip_resp[strip_num, :], alpha=0.6)
    plt.xlabel("Time (ms)", fontsize=fontsize_value)
    plt.ylabel("Strip response (AU)", fontsize=fontsize_value)
    plt.tick_params(axis="both", which="major", labelsize=10)
    plt.show()

# noize calculation on first 100 events
noize = [np.mean(raw_strip_resp[strip_num, 0:100]) for strip_num in range(18, 153)]
noize = np.concatenate((np.zeros(18), noize))
if fill_excel_data:
    fill_excel(
        res_excel_file, "mb_scan_ESRF", [strip for strip in range(18, 153)], 4, 3
    )
    fill_excel(res_excel_file, "mb_scan_ESRF", noize[18:153], 4, 4)

# cut noize
strip_noize = np.repeat(noize, len(time_values)).reshape(-1, len(time_values))
strip_resp = raw_strip_resp - strip_noize

# get normalization values of each strip
normal_val = calc_normal_val(time_values, strip_resp)
if fill_excel_data:
    fill_excel(res_excel_file, "mb_scan_ESRF", normal_val[18:153], 4, 5)

# normalize strip resp
normal_val = np.repeat(normal_val, len(time_values)).reshape(-1, len(time_values))
normal_strip_resp = strip_resp / normal_val

# normal strip resp plot
if plot_strip_resp:
    for strip_num in range(18, 153):
        plt.plot(time_values, normal_strip_resp[strip_num, :], alpha=0.6)
    plt.xlabel("Time (ms)", fontsize=fontsize_value)
    plt.ylabel("Normalized strip response (AU)", fontsize=fontsize_value)
    plt.tick_params(axis="both", which="major", labelsize=10)
    plt.show()
