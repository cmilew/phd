import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.signal import find_peaks, peak_widths
from scipy.interpolate import interp1d
import sys


def read_data(data_file, correspondence_table_file):
    """Function to read data from measurement .bin file"""
    # .bin measurements file
    dt = np.dtype("uint16")
    f = open(data_file, "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)

    # correspondence of QDC number and strip number file
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()
    # number of measurements
    nb_events = np.size(zdata) // 309

    # time conversion in seconds (integration time = 0.01s + 0.0005 s of dead time)
    time_values = np.array([event * 0.0105 for event in range(nb_events)])

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, len(time_values)))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_events):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )

    return time_values, raw_strip_resp


def calc_center_peak_resp(time, resp):
    """Function calculating the normalization value of each strip which corresponds to
    the center of the peak response"""

    center_peak_resp = np.zeros(153)
    time_center_peak = np.zeros(153)

    for strip in range(18, 153):
        # find peak in strip resp
        peak_index, _ = find_peaks(resp[strip], height=2500)
        assert len(peak_index) == 1, (
            "None or more than one peak found for strip {}".format(strip)
        )

        # find FWHM
        fwhm = peak_widths(resp[strip], peak_index, rel_height=0.5)
        fwhm_left = np.interp(fwhm[2][0], np.arange(len(time)), time)
        fwhm_right = np.interp(fwhm[3][0], np.arange(len(time)), time)

        # calc coord of mid FWHM = center value of peak
        x_mid_peak = (fwhm_left + fwhm_right) / 2
        time_center_peak[strip] = x_mid_peak
        y_mid_peak = np.interp(x_mid_peak, time_values, strip_resp[strip])

        center_peak_resp[strip] = y_mid_peak

        # reformat center_peak_resp to have the same shape as strip resp to normalize this array with it
        center_peak_resp_reformat = np.repeat(center_peak_resp, len(time)).reshape(
            -1, len(time)
        )

    return time_center_peak, center_peak_resp_reformat


def calc_mean_speed(time_values, raw_strip_resp):
    """Function calculating the lateral speed of scanning for every diamond detector by
    dividing the distance between the second and penultimate strip of each diamond by
    by the time between the response peaks of these strips.
    Returns the mean speed overall."""

    # calc between second and penultimate strip of each diamond
    # strip 17 first diamond not functional so starting at 19
    start_end_strips = [
        [32, 19],
        [49, 35],
        [66, 52],
        [83, 69],
        [100, 86],
        [117, 103],
        [134, 120],
        [151, 137],
    ]

    speeds = []
    for idiamond in range(0, 8):
        start_end_times = []
        for strip in start_end_strips[idiamond]:
            peak_index, _ = find_peaks(raw_strip_resp[strip], height=2500)
            assert len(peak_index) == 1, (
                "No peak or more than one peak found for {}".format(strip)
            )
            start_end_times.append(time_values[peak_index[0]])

        # dist between first and last strip of a diamond detector (on first diamond
        # only 16 strips responding)
        if idiamond == 0:
            det_width = 13 * 0.2325
        else:
            det_width = 14 * 0.2325  # mm
        speed_diamond = det_width / (start_end_times[1] - start_end_times[0])  # mm / s
        speeds.append(speed_diamond)
    speeds = np.array(speeds)
    mean_speed = np.mean(speeds)
    return mean_speed


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l + i, column=start_col, value=data_to_fill[i])
    wb_res.save(excel_path)


#### TO FILL ####
fontsize_value = 20
plot_raw_resp = True
plot_strip_resp = True

# mid strip = 77 (= 153 / 2) + 17 first strip of missing diamond  - 1 (index start at 0)
mid_strip = 77 + 17 - 1
fill_excel_bool = True
bool_pos_fn_time = False


# read measurements data
mes_file = r"C:\Users\milewski\Desktop/these/mesures/caracterisation_detecteur_8-06-2023/more_homogeneous_zData_150V_1ubeam_24p8v0_scan6.bin"
correspondence_table_file = r"C:\Users\milewski\Desktop/these/mesures/analyse_data/codes_python/150_voies/add_piste.txt"
res_file = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"
position_excel_file = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\microbeam_scan_analysis\strip_exact_positioning.xlsx"

time_values, raw_strip_resp = read_data(mes_file, correspondence_table_file)


# noize calculation on first 100 events
noize = np.mean(raw_strip_resp[:, 0:100], axis=1)

# cut noize
strip_noize = np.repeat(noize, len(time_values)).reshape(-1, len(time_values))
strip_resp = raw_strip_resp - strip_noize

# normalize strip resp
time_center_peak, center_peak_resp = calc_center_peak_resp(time_values, strip_resp)
normal_strip_resp = strip_resp[17:153] / center_peak_resp[17:153]
normal_strip_resp = np.concatenate(
    (np.zeros((17, len(time_values))), normal_strip_resp)
)


# convert time values in position
mean_speed = calc_mean_speed(time_values, raw_strip_resp)
positions = time_values * mean_speed
peaks_positions = time_center_peak * mean_speed

# center positions on mid strip
centered_positions = positions - peaks_positions[mid_strip]

peaks_positions_centered = peaks_positions - peaks_positions[mid_strip]


# saves centered positions in excel file
if fill_excel_bool:
    fill_excel(position_excel_file, "test", peaks_positions, 3, 4)

# calculates total sum of responses
sum_resp = np.sum(raw_strip_resp, axis=0)
normal_sum_resp = np.nansum(normal_strip_resp, axis=0)


# raw strips response plot
if plot_raw_resp:
    for strip_num in range(18, 153):
        plt.plot(
            centered_positions,
            raw_strip_resp[strip_num, :],
            alpha=0.6,
            label=f"Strip {strip_num}",
        )
        plt.plot(
            centered_positions,
            sum_resp,
            label="Strip response sum",
            color="black",
            linewidth=1.2,
            linestyle="-",
        )
    plt.xlabel("Positions (mm)", fontsize=fontsize_value)
    plt.ylabel("Raw strip response (AU)", fontsize=fontsize_value)
    plt.tick_params(axis="both", which="major", labelsize=fontsize_value)
    plt.xlim(-23, 20)
    # plt.legend()
    plt.show()


# normal strip resp plot
if plot_strip_resp:
    for strip_num in range(18, 153):
        plt.plot(centered_positions, normal_strip_resp[strip_num, :], alpha=0.6)
    plt.plot(
        centered_positions,
        normal_sum_resp,
        label="Strip response sum",
        color="black",
        linewidth=1.2,
        linestyle="-",
    )
    plt.xlabel("Positions (mm)", fontsize=fontsize_value)
    # plt.title("Lateral scan with one microbeam", fontsize=fontsize_value)
    plt.xlim(-23, 20)
    plt.ylabel("Normalized strip response (AU)", fontsize=fontsize_value)
    plt.tick_params(axis="both", which="major", labelsize=fontsize_value)
    # plt.legend()
    plt.show()

if bool_pos_fn_time:
    plt.scatter(
        time_center_peak[time_center_peak != 0], peaks_positions[peaks_positions != 0]
    )
    plt.xlabel("Time (s)", fontsize=fontsize_value)
    plt.ylabel("Position (mm)", fontsize=fontsize_value)
    plt.show()
