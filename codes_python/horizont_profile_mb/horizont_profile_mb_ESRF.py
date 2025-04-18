import matplotlib.pyplot as plt
import numpy as np
import sys
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score
import matplotlib.ticker as ticker
import math


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def get_excel_data(excel_file, ws_name, start_l, start_col):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(excel_file)
    ws = wb_res[ws_name]
    excel_data = [ws.cell(row=start_l + i, column=start_col).value for i in range(135)]
    excel_data = np.concatenate((np.zeros(18), excel_data))
    return excel_data


def get_and_format_noize_normal_val(
    excel_file, ws_name, start_l, start_col, time_values
):
    noize = get_excel_data(excel_file, ws_name, start_l, start_col)
    normal_val = get_excel_data(excel_file, ws_name, start_l, start_col + 1)
    noize = np.repeat(noize, len(time_values)).reshape(-1, len(time_values))
    noize_strips = np.concatenate((np.zeros((18, len(time_values))), noize))
    normal_val = np.repeat(normal_val, len(time_values)).reshape(-1, len(time_values))
    normal_val_strips = np.concatenate((np.zeros((18, len(time_values))), normal_val))

    return noize_strips, normal_val_strips


def calculates_mean_value_plateau(
    time_values, start_time_value, nb_of_averaged_points, strip_resp
):
    """function calculating the mean value of strip responses over a range of number of
    time values given in argument (with starting point), normally on a plateau of the
    step phantom. Returns an array of mean values for each strip."""

    start_event_index = find_closest_index(time_values, start_time_value)
    last_event_index = int(start_event_index + nb_of_averaged_points)
    mean_values_plateau = np.zeros(len(strip_resp))
    std_val_plateau = np.zeros(len(strip_resp))
    for strip in range(0, len(strip_resp)):
        mean_values_plateau[strip] = np.mean(
            strip_resp[strip, start_event_index:last_event_index]
        )
        std_val_plateau[strip] = np.std(
            strip_resp[strip, start_event_index:last_event_index]
        )
    return mean_values_plateau, std_val_plateau


def read_bin_file():
    # correspondence of QDC number and strip number file
    correspondence_table_file = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_mes = np.size(zdata) // 309

    # time conversion (integration time = 10 ms + 0.5 ms of dead time)
    time_values = [event * 10.5 for event in range(nb_mes)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, nb_mes))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_mes):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )
    return time_values, raw_strip_resp


def sort_mes_by_strip_pos(
    strip_resp, strip_positions, intermb_thresh, off_p_thresh, std_strip_resp=None
):
    """Function sorting strip positions and measured values by their position :
    interbeam, off pitch and facing microbeam"""

    intermb_strip_pos = []
    intermb_resp = []
    std_intermb = []
    mb_strip_pos = []
    mb_resp = []
    std_mb = []
    off_p_strip_pos = []
    off_p_resp = []
    std_off_p = []

    for iresp, resp in enumerate(strip_resp):
        if resp < intermb_thresh:
            intermb_strip_pos.append(strip_positions[iresp])
            intermb_resp.append(strip_resp[iresp])
            if std_strip_resp is not None:
                std_intermb.append(std_strip_resp[iresp])
        elif resp > intermb_thresh and resp < off_p_thresh:
            off_p_resp.append(strip_resp[iresp])
            off_p_strip_pos.append(strip_positions[iresp])
            if std_strip_resp is not None:
                std_off_p.append(std_strip_resp[iresp])
        else:
            mb_resp.append(strip_resp[iresp])
            mb_strip_pos.append(strip_positions[iresp])
            if std_strip_resp is not None:
                std_mb.append(std_strip_resp[iresp])
    return (
        intermb_strip_pos,
        intermb_resp,
        std_intermb,
        mb_strip_pos,
        mb_resp,
        std_mb,
        off_p_strip_pos,
        off_p_resp,
        std_off_p,
    )


# .bin measurements file ESRF
name_mes_file = r"\zData_150V_150ubeam_24p8v0_40step8_0cmvitesse10.bin"
file_mes_path = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\mesures_carac_det_8-06-2023\step_phantom\step_0_8cm"
mes_file = file_mes_path + name_mes_file
dt = np.dtype("uint16")
f = open(mes_file, "rb")
data = f.read()
zdata = np.frombuffer(data, dt)

# Parameter file
param_file = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"

# Gets parameters values for analysis
df_param = pd.read_excel(param_file, sheet_name="param_analyse", skiprows=2)
param = df_param["ESRF"].to_numpy()

# Nb of steps in step phantom
nb_of_steps = int(param[0])
step_thicknesses = [0, 1, 2, 3, 4, 6]

# Number of points for noize calculation
nb_points_noize = int(param[1])

# Start time for normalization value calculation
start_time_normal_calc = param[2]
nb_points_normal_calc = param[3]

# Start time values used for averaging of each phantom steps
start_time_values_each_thicknesses = [
    time_val for time_val in param[4 : 4 + nb_of_steps]
]
nb_points_mean_plateau_calc = int(param[10])

# Get data from measurements file
time_values, raw_strip_resp = read_bin_file()

# Calculate mean of measures on step 0 plateau
resp_step_0, std_step_0 = calculates_mean_value_plateau(
    time_values, start_time_normal_calc, nb_points_normal_calc, raw_strip_resp
)

# Get noize, normal values and strip positions from microbeam lateral scan
noize = get_excel_data(param_file, "mb_scan_ESRF", 4, 4)
normal_val = get_excel_data(param_file, "mb_scan_ESRF", 4, 5)
strip_positions = get_excel_data(param_file, "mb_scan_ESRF", 4, 6)

font_size = 15

# # Sort data by strip position : facing mb, inter mb of off pitch
# (
#     intermb_strip_pos,
#     intermb_resp,
#     std_intermb,
#     mb_strip_pos,
#     mb_resp,
#     std_mb,
#     off_p_strip_pos,
#     off_p_resp,
#     std_off_p,
# ) = sort_mes_by_strip_pos(resp_step_0, strip_positions, 5000, 15000, std_step_0)

# Sort data by even or odd number
even_strips_resp = []
even_strips_pos = []
odd_strips_resp = []
odd_strips_pos = []
for strip in range(18, 153):
    if strip % 2 == 0:
        even_strips_resp.append(resp_step_0[strip])
        even_strips_pos.append(strip_positions[strip])
    else:
        odd_strips_resp.append(resp_step_0[strip])
        odd_strips_pos.append(strip_positions[strip])

print(f"odd_strips_resp[0] = {odd_strips_resp[0]}")

# Plot raw data on step 0
plt.scatter(
    even_strips_pos,
    even_strips_resp,
    color="orange",
    label="even strips",
)
plt.scatter(odd_strips_pos, odd_strips_resp, color="blue", label="odd strips")
# plt.scatter(intermb_strip_pos, intermb_resp, label="Interbeam strips", color="blue")
# plt.errorbar(
#     intermb_strip_pos, intermb_resp, yerr=std_intermb, linestyle="None", capsize=5
# )
# plt.scatter(mb_strip_pos, mb_resp, label="Microbeam strips", color="orange")
# plt.errorbar(mb_strip_pos, mb_resp, yerr=std_mb, linestyle="None", capsize=5)
# plt.scatter(off_p_strip_pos, off_p_resp, label="Off pitch strips", color="green")
# plt.errorbar(off_p_strip_pos, off_p_resp, yerr=std_off_p, linestyle="None", capsize=5)
plt.xlabel("Strip position (mm)", fontsize=font_size)
plt.ylabel("Strip response (AU)", fontsize=font_size)
y_ticks = plt.yticks()[0]
new_y_ticks = np.append(y_ticks, 500)
plt.yticks(new_y_ticks)
plt.ylim(0, 25000)
plt.title("Strip responses MRT", fontsize=font_size)
plt.legend(fontsize=font_size)
plt.show()

# PVDR
if len(even_strips_resp) < len(odd_strips_resp):
    len_pvdr = len(even_strips_resp)
    pvdr_pos = even_strips_pos
else:
    len_pvdr = len(odd_strips_resp)
    pvdr_pos = odd_strips_pos
pvdr = np.array(
    [
        even_strips_resp[i] / odd_strips_resp[i]
        if even_strips_resp[i] > odd_strips_resp[i]
        else odd_strips_resp[i] / even_strips_resp[i]
        for i in range(len_pvdr)
    ]
)

# Retrieves only 40 first values (loss of pitch after)
pvdr_pos = pvdr_pos[:40]
pvdr = pvdr[:40]

# Plot PVDR
plt.figure()
plt.scatter(pvdr_pos, pvdr, color="black")
plt.xlim(-15, 6)
plt.ylim(0, 50)
plt.xlabel("Position (mm)", fontsize=font_size)
plt.ylabel("In beam response / interbeam response", fontsize=font_size)
plt.title("In beam inter beam response ratio ESRF", fontsize=font_size)
plt.show()


# Cut noize and normalise mean of step 0 plateau
normal_resp_step_0 = (resp_step_0 - noize) / normal_val

# # Sort normalized resp
# (
#     normal_intermb_pos,
#     normal_intermb_resp,
#     normal_std_intermb,
#     normal_mb_pos,
#     normal_mb_resp,
#     normal_std_mb,
#     normal_off_p_pos,
#     normal_off_p_resp,
#     normal_std_off_p,
# ) = sort_mes_by_strip_pos(normal_resp_step_0, strip_positions, 0.2, 0.8)


# Sort data by even or odd number
normal_even_strips_resp = []
normal_odd_strips_resp = []
for strip in range(18, 153):
    if strip % 2 == 0:
        normal_even_strips_resp.append(normal_resp_step_0[strip])
    else:
        normal_odd_strips_resp.append(normal_resp_step_0[strip])


# Plot normal data on step 0
plt.scatter(
    even_strips_pos,
    normal_even_strips_resp,
    color="orange",
    label="even strips",
)
plt.scatter(odd_strips_pos, normal_odd_strips_resp, color="blue", label="odd strips")
# plt.scatter(
#     normal_intermb_pos, normal_intermb_resp, label="Interbeam strips", color="blue"
# )
# plt.scatter(normal_mb_pos, normal_mb_resp, label="Microbeam strips", color="orange")
# plt.scatter(
#     normal_off_p_pos, normal_off_p_resp, label="Off pitch strips", color="green"
# )
y_ticks = plt.yticks()[0]
new_y_ticks = np.append(y_ticks, 0.02)
new_y_ticks = np.delete(new_y_ticks, 0)
plt.yticks(new_y_ticks)
# plt.ylim(-1, 1.5)
plt.xlabel("Strip position (mm)", fontsize=font_size)
plt.ylabel("Strip response (AU)", fontsize=font_size)
plt.title("Normalized strip responses MRT", fontsize=font_size)
plt.legend(fontsize=font_size)
plt.show()
