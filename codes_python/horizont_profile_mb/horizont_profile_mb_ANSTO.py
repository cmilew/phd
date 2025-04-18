import matplotlib.pyplot as plt
import numpy as np
import sys
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score
import matplotlib.ticker as ticker
import math


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def get_excel_data(excel_file, ws_name, start_l, start_col):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(excel_file)
    ws = wb_res[ws_name]
    excel_data = [ws.cell(row=start_l + i, column=start_col).value for i in range(135)]
    excel_data = np.concatenate((np.zeros(18), excel_data))
    return excel_data


def get_and_format_noize_normal_val(
    excel_file, ws_name, start_l, start_col, time_values
):
    noize = get_excel_data(excel_file, ws_name, start_l, start_col)
    normal_val = get_excel_data(excel_file, ws_name, start_l, start_col + 1)
    noize = np.repeat(noize, len(time_values)).reshape(-1, len(time_values))
    noize_strips = np.concatenate((np.zeros((18, len(time_values))), noize))
    normal_val = np.repeat(normal_val, len(time_values)).reshape(-1, len(time_values))
    normal_val_strips = np.concatenate((np.zeros((18, len(time_values))), normal_val))

    return noize_strips, normal_val_strips


def calculates_mean_value_plateau(
    time_values, start_time_value, nb_of_averaged_points, strip_resp
):
    """function calculating the mean value of strip responses over a range of number of
    time values given in argument (with starting point), normally on a plateau of the
    step phantom. Returns an array of mean values for each strip."""

    start_event_index = find_closest_index(time_values, start_time_value)
    last_event_index = int(start_event_index + nb_of_averaged_points)
    mean_values_plateau = np.zeros(len(strip_resp))
    std_val_plateau = np.zeros(len(strip_resp))
    for strip in range(0, len(strip_resp)):
        mean_values_plateau[strip] = np.mean(
            strip_resp[strip, start_event_index:last_event_index]
        )
        std_val_plateau[strip] = np.std(
            strip_resp[strip, start_event_index:last_event_index]
        )
    return mean_values_plateau, std_val_plateau


def read_bin_file():
    # correspondence of QDC number and strip number file
    correspondence_table_file = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
    dt = np.dtype("uint16")
    f = open(mes_file, "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_mes = np.size(zdata) // 309

    # time conversion (integration time = 10 ms + 0.5 ms of dead time)
    time_values = [event * 10.5 for event in range(nb_mes)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, nb_mes))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_mes):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )
    return time_values, raw_strip_resp


def sort_mes_by_strip_pos(
    strip_resp, strip_positions, intermb_thresh, off_p_thresh, std_strip_resp=None
):
    """Function sorting strip positions and measured values by their position :
    interbeam, off pitch and facing microbeam"""

    intermb_strip_pos = []
    intermb_resp = []
    std_intermb = []
    mb_strip_pos = []
    mb_resp = []
    std_mb = []
    off_p_strip_pos = []
    off_p_resp = []
    std_off_p = []

    for iresp, resp in enumerate(strip_resp):
        if resp < intermb_thresh:
            intermb_strip_pos.append(strip_positions[iresp])
            intermb_resp.append(strip_resp[iresp])
            if std_strip_resp is not None:
                std_intermb.append(std_strip_resp[iresp])
        elif resp > intermb_thresh and resp < off_p_thresh:
            off_p_resp.append(strip_resp[iresp])
            off_p_strip_pos.append(strip_positions[iresp])
            if std_strip_resp is not None:
                std_off_p.append(std_strip_resp[iresp])
        else:
            mb_resp.append(strip_resp[iresp])
            mb_strip_pos.append(strip_positions[iresp])
            if std_strip_resp is not None:
                std_mb.append(std_strip_resp[iresp])
    return (
        intermb_strip_pos,
        intermb_resp,
        std_intermb,
        mb_strip_pos,
        mb_resp,
        std_mb,
        off_p_strip_pos,
        off_p_resp,
        std_off_p,
    )


# .bin measurements file ESRF
name_mes_file = r"\4T_Al_Cu_step_fant_1_2_2.5_cm_VREF_UC_0_sans_colli_MRT.bin"
file_mes_path = r"C:\Users\milewski\Desktop\these\mission_australie\mesures\136_voies\20-03-2024\fantome_escalier\AlCu\MRT"
mes_file = file_mes_path + name_mes_file
font_size = 15
beam_type = "AlCu"


# Parameter file
param_file = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"

# Gets parameters values for analysis
df_param = pd.read_excel(param_file, sheet_name="param_analyse", skiprows=2)
col_name = beam_type + " 0-1-2-2.5"
param = df_param[col_name].to_numpy()

# Number of points for noize calculation
nb_points_noize = int(param[1])

# Start time for normalization value calculation
start_time_normal_calc = param[2]
nb_points_normal_calc = param[3]

# Get data from measurements file
time_values, raw_strip_resp = read_bin_file()

# Calculate mean of measures on step 0 plateau
resp_step_0, std_step_0 = calculates_mean_value_plateau(
    time_values, start_time_normal_calc, nb_points_normal_calc, raw_strip_resp
)

# Calc noize and get strip positions from microbeam lateral scan at ESRF
noize = np.mean(raw_strip_resp[:, :nb_points_noize], axis=1)
strip_positions = get_excel_data(param_file, "mb_scan_ESRF", 4, 6)


# Plot raw data on step 0
plt.scatter(strip_positions, resp_step_0, color="blue")
plt.xlabel("Strip position (mm)", fontsize=font_size)
plt.ylabel("Strip response (AU)", fontsize=font_size)
# plt.ylim(-100, 12000)
y_ticks = plt.yticks()[0]
new_y_ticks = np.append(y_ticks, 1000)
plt.yticks(new_y_ticks)
plt.title(f"Strip responses ANSTO poly {beam_type}", fontsize=font_size)
plt.legend(fontsize=font_size)
plt.show()

# Sort data by even or odd number
even_strips_resp = []
even_strips_pos = []
odd_strips_resp = []
odd_strips_pos = []
for strip in range(18, 153):
    if strip % 2 == 0:
        even_strips_resp.append(resp_step_0[strip])
        even_strips_pos.append(strip_positions[strip])
    else:
        odd_strips_resp.append(resp_step_0[strip])
        odd_strips_pos.append(strip_positions[strip])
for strip in range(60, 80):
    print(f"strip {strip} resp: {resp_step_0[strip]}")
# Plot raw data on step 0
plt.scatter(even_strips_pos, even_strips_resp, color="blue", label="Even strips")
plt.scatter(odd_strips_pos, odd_strips_resp, color="orange", label="Odd strips")
plt.xlabel("Strip position (mm)", fontsize=font_size)
plt.ylabel("Strip response (AU)", fontsize=font_size)
# plt.ylim(-100, 12000)
y_ticks = plt.yticks()[0]
new_y_ticks = np.append(y_ticks, 1000)
plt.yticks(new_y_ticks)
plt.title(f"Strip responses ANSTO poly {beam_type}", fontsize=font_size)
plt.legend(fontsize=font_size)
plt.show()

# PVDR
if len(even_strips_resp) < len(odd_strips_resp):
    len_pvdr = len(even_strips_resp)
    pvdr_pos = even_strips_pos
else:
    len_pvdr = len(odd_strips_resp)
    pvdr_pos = odd_strips_pos
pvdr = np.array(
    [
        even_strips_resp[i] / odd_strips_resp[i]
        if even_strips_resp[i] > odd_strips_resp[i]
        else odd_strips_resp[i] / even_strips_resp[i]
        for i in range(len_pvdr)
    ]
)

# Plot PVDR
plt.figure()
plt.scatter(pvdr_pos, pvdr, color="black")
plt.xlabel("PVDR position (mm)", fontsize=font_size)
plt.ylabel("PVDR", fontsize=font_size)
plt.title(f"PVDR {beam_type}", fontsize=font_size)
plt.show()

sys.exit()
# Cut noize and normalise mean of step 0 plateau
normal_resp_step_0 = (resp_step_0 - noize) / normal_val

# Sort normalized resp
(
    normal_intermb_pos,
    normal_intermb_resp,
    normal_std_intermb,
    normal_mb_pos,
    normal_mb_resp,
    normal_std_mb,
    normal_off_p_pos,
    normal_off_p_resp,
    normal_std_off_p,
) = sort_mes_by_strip_pos(normal_resp_step_0, strip_positions, 0.2, 0.8)


# Plot normal data on step 0
plt.scatter(
    normal_intermb_pos, normal_intermb_resp, label="Interbeam strips", color="blue"
)
plt.scatter(normal_mb_pos, normal_mb_resp, label="Microbeam strips", color="orange")
plt.scatter(
    normal_off_p_pos, normal_off_p_resp, label="Off pitch strips", color="green"
)
plt.xlabel("Strip position (mm)", fontsize=font_size)
plt.ylabel("Strip response (AU)", fontsize=font_size)
plt.title("Normalized strip responses MRT", fontsize=font_size)
plt.legend(fontsize=font_size)
plt.show()
