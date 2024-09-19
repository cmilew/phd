import matplotlib.pyplot as plt
import numpy as np
import sys
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score
import matplotlib.ticker as ticker
import math


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def exponential_model(x, a, b):
    return a * np.exp(-b * x)


def cut_noize_and_normalize(raw_strip_resp, strip_noize, normalization_value, factor):
    clean_data_fn = (
        lambda x: (x - strip_noize) / (normalization_value - strip_noize) * factor
    )
    vectorized_clean_data_fn = np.vectorize(clean_data_fn)
    normalized_strip_resp = vectorized_clean_data_fn(raw_strip_resp)
    return normalized_strip_resp


def calculates_mean_value_plateau(
    time_values, start_time_value, nb_of_averaged_points, strip_resp
):
    """function calculating the mean value of strip responses over a range of number of
    time values given in argument (with starting point), normally on a plateau of the
    step phantom. Returns an array of mean values for each strip."""

    start_event_index = find_closest_index(time_values, start_time_value)
    last_event_index = int(start_event_index + nb_of_averaged_points)
    mean_values_plateau = np.zeros(len(strip_resp))
    std_val_plateau = np.zeros(len(strip_resp))
    for strip in range(0, len(strip_resp)):
        mean_values_plateau[strip] = np.mean(
            strip_resp[strip, start_event_index:last_event_index]
        )
        std_val_plateau[strip] = np.std(
            strip_resp[strip, start_event_index:last_event_index]
        )
    return mean_values_plateau, std_val_plateau


def remove_strip_from_list(
    list_of_strips, strip_to_rem, current_file_name, file_name_strip_to_rem
):
    if current_file_name == file_name_strip_to_rem:
        for strip in strip_to_rem:
            list_of_strips.remove(strip)
    return list_of_strips


def fill_mu_in_excel(ws_mu, strips_list, mass_att_coeff_list, start_l, start_col):
    for i in range(len(mass_att_coeff_list)):
        ws_mu.cell(row=start_l + i, column=start_col, value=strips_list[i])
        ws_mu.cell(row=start_l + i, column=start_col + 1, value=mass_att_coeff_list[i])
    wb_res.save(param_res_file)


def gaussian(x, amplitude, mean, stddev):
    """Gaussian function"""
    return amplitude * np.exp(-((x - mean) ** 2) / (2 * stddev**2))


def plot_hist_gaussian_fit(mu_values, strips_region):
    # hist, bin_edges = np.histogram(mu_values, bins=5, density=True)
    hist, bin_edges = np.histogram(mu_values, bins=10)

    # bin centers def for gaussian fit
    bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2

    # Gaussian fit
    popt, pcov = curve_fit(
        gaussian,
        bin_centers,
        hist,
        p0=[max(hist), np.mean(mu_values), np.std(mu_values)],
        maxfev=5000,
    )

    amplitude_gaussian_fit, mean_gaussian_fit, stddev_gaussian_fit = popt
    x_gaussian_data = np.linspace(
        mean_gaussian_fit - 2 * stddev_gaussian_fit - 0.5 * stddev_gaussian_fit,
        mean_gaussian_fit + 2 * stddev_gaussian_fit + 0.5 * stddev_gaussian_fit,
        100,
    )

    # plt.hist(mu_values, bins=10, density=True, color="#173896", edgecolor="white")
    plt.hist(mu_values, bins=10, color="#173896", edgecolor="white")
    plt.plot(
        x_gaussian_data,
        gaussian(x_gaussian_data, *popt),
        "-",
        color="orange",
        label="Gaussian fit",
        linewidth=line_width,
    )
    plt.axvline(
        mean_gaussian_fit,
        color="red",
        linestyle="--",
        label="Mean",
        linewidth=line_width,
    )
    plt.axvline(
        mean_gaussian_fit - 2 * stddev_gaussian_fit,
        color="green",
        linestyle="--",
        label="±2σ",
        linewidth=line_width,
    )
    plt.axvline(
        mean_gaussian_fit + 2 * stddev_gaussian_fit,
        color="green",
        linestyle="--",
        linewidth=line_width,
    )

    # format axis values
    formatter = ticker.FormatStrFormatter("%.4f")
    plt.gca().xaxis.set_major_formatter(formatter)
    plt.title("Mass attenuation coeff " + strips_region)
    plt.xlabel("Mass attenuation coefficient (cm-1)")
    plt.ylabel("Number of strips")
    plt.legend()
    plt.show()


def get_med_percent(data):
    median = np.median(data)
    q1 = np.percentile(data, 25)
    q3 = np.percentile(data, 75)
    return median, q1, q3


def read_bin_file():
    # correspondence of QDC number and strip number file
    correspondence_table_file = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_mes = np.size(zdata) // 309

    # time conversion (integration time = 10 ms + 0.5 ms of dead time)
    time_values = [event * 10.5 for event in range(nb_mes)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, nb_mes))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_mes):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )
    return time_values, raw_strip_resp


# .bin measurements file ESRF
name_mes_file = r"\zData_150V_150ubeam_24p8v0_40step8_0cmvitesse10.bin"
file_mes_path = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\mesures_carac_det_8-06-2023\step_phantom\step_0_8cm"
mes_file = file_mes_path + name_mes_file
dt = np.dtype("uint16")
f = open(mes_file, "rb")
data = f.read()
zdata = np.frombuffer(data, dt)

# Parameter and results file
param_res_file = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"
wb_res = load_workbook(param_res_file)
df_res = pd.read_excel(param_res_file, sheet_name="mu_ESRF", skiprows=2)

###### PARAM TO ADJUST ######
beam_type = "ESRF"
ws_mu = wb_res["mu_ESRF"]
ws_normal_val = wb_res["mes_ESRF"]
line_width = 2
# Line where to start to write mu results
start_l = 4
start_col = 3
##############################

# Gets parameters values for analysis
df_param = pd.read_excel(param_res_file, sheet_name="param_analyse", skiprows=2)
param = df_param[beam_type].to_numpy()

# Nb of steps in step phantom
nb_of_steps = int(param[0])
step_thicknesses = [0, 1, 2, 3, 4, 6]

# Number of points for noize calculation
nb_points_noize = int(param[1])

# Start time for normalization value calculation
start_time_normal_calc = param[2]
nb_points_normal_calc = param[3]

# Start time values used for averaging of each phantom steps
start_time_values_each_thicknesses = [
    time_val for time_val in param[4 : 4 + nb_of_steps]
]
nb_points_mean_plateau_calc = int(param[10])

time_values, raw_strip_resp = read_bin_file()

strip_normal_val, std_normal_values = calculates_mean_value_plateau(
    time_values, start_time_normal_calc, nb_points_normal_calc, raw_strip_resp
)

# fill_excel(
#     wb_res,
#     ws_normal_val,
#     param_res_file,
#     [strip for strip in range(18, 153)],
#     start_l,
#     start_col,
# )
# fill_excel(
#     wb_res,
#     ws_normal_val,
#     param_res_file,
#     strip_normal_val[18:153],
#     start_l,
#     start_col + 1,
# )

# # raw strips response plot
# plt.figure()
# for strip_num in range(18, 153):
#     plt.plot(time_values, raw_strip_resp[strip_num, :], label=str(strip_num))
# plt.xlabel("Time (ms)")
# plt.ylabel("QDC response (AU)")
# plt.title(f"Raw responses ESRF", fontsize=18, fontstyle="oblique")
# plt.show()

# noize calculation
strip_noizes = [
    np.mean(raw_strip_resp[strip, 0:nb_points_noize]) for strip in range(18, 153)
]
strip_noizes = np.concatenate((np.zeros(18), strip_noizes))

# fill_excel(
#     wb_res, ws_normal_val, param_res_file, strip_noizes[18:153], start_l, start_col + 2
# )
# sys.exit()
# # noize cut and normalization
# normalized_strip_resp = np.zeros((153, len(time_values)))
# for strip in range(18, 153):
#     normalized_strip_resp[strip, :] = cut_noize_and_normalize(
#         raw_strip_resp[strip],
#         strip_noizes[strip],
#         strip_normal_val[strip],
#         1,
#     )


# List with each row corresponding to a strip and each column corresponding to the
# mean resp of this strip for each step thickness
# mean_strip_resp = []
# std_mean_strip_resp = []
# for i in range(nb_of_steps):
#     mean_val, std_val = calculates_mean_value_plateau(
#         time_values,
#         start_time_values_each_thicknesses[i],
#         nb_points_mean_plateau_calc,
#         normalized_strip_resp,
#     )
#     mean_strip_resp.append(mean_val)
#     std_mean_strip_resp.append(std_val)

# # exponential fit to retrieve mass att coeff
# mass_att_coeff = np.zeros(153)
# for strip in range(18, 153):
#     popt, pcov = curve_fit(
#         exponential_model,
#         step_thicknesses,
#         [mean_strip_resp[step][strip] for step in range(len(step_thicknesses))],
#         p0=(1, 1),
#     )
#     a_opt, b_opt = popt
#     mass_att_coeff[strip] = b_opt


# # separates strip resp between region : facing microbeams, interbeams and off pitch
# mb_strips = [
#     strip
#     for strip in range(18, 153)
#     if low_value_mb < strip_normal_val[strip]
# ]
# mb_mu = [mass_att_coeff[strip] for strip in mb_strips]

# interb_strips = [
#     strip
#     for strip in range(18, 153)
#     if low_value_inter_b < strip_normal_val[strip] < low_value_off_p
# ]
# interb_mu = [mass_att_coeff[strip] for strip in interb_strips]

# off_p_strips = [
#     strip
#     for strip in range(18, 153)
#     if low_value_off_p < strip_normal_val[strip] < low_value_mb
# ]
# off_p_mu = [mass_att_coeff[strip] for strip in off_p_strips]

# fills excel result file
# fill_mu_in_excel(ws_mu, mb_strips, mb_mu, start_l_res, start_col_res)
# fill_mu_in_excel(ws_mu, interb_strips, interb_mu, start_l_res, start_col_res + 3)
# fill_mu_in_excel(ws_mu, off_p_strips, off_p_mu, start_l_res, start_col_res + 6)

# # Gaussian distribution fit
# plot_hist_gaussian_fit(mb_mu, "microbeam strips")
# plot_hist_gaussian_fit(interb_mu, "interbeam strips")
plt.hist(mu_interb, bins=10, color="#173896", edgecolor="white")
plt.title("Mass attenuation coeff interbeam strips")
plt.xlabel("Mass attenuation coefficient (cm-1)")
plt.ylabel("Number of strips")
plt.show()
# plot_hist_gaussian_fit(off_p_mu_cut, "off pitch strips")

mu_off_p = [mu for mu in mu_off_p if not math.isnan(mu)]

# boxplot
fig, ax = plt.subplots(figsize=(8, 6))
ax.boxplot(mu_mb, patch_artist=True, vert=True, widths=0.5)
plt.show()
sys.exit()
ax.boxplot([mu_mb, mu_interb, mu_off_p], patch_artist=True, vert=True, widths=0.5)

med_mb, q1_mb, q3_mb = get_med_percent(mu_mb)
med_interb, q1_interb, q3_interb = get_med_percent(mu_interb)
med_off_p, q1_off_p, q3_off_p = get_med_percent(mu_off_p)

ax.set_xticklabels(["Strips facing microbeams", "Interbeam strips", "Off pitch strips"])
ticks = np.arange(0.08, 0.18, 0.005)
ax.set_yticks(ticks)
ax.legend()
plt.title("Linear attenuation coefficient distribution")
plt.ylabel("Linear attenuation coefficient (cm-1)")
plt.grid(True)

plt.show()
