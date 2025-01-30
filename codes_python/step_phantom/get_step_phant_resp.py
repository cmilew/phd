import matplotlib.pyplot as plt
import numpy as np
import sys
import pandas as pd
from openpyxl import load_workbook


def read_bin_file(zdata):
    # correspondence of QDC number and strip number file
    correspondence_table_file = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_mes = np.size(zdata) // 309

    # time conversion (integration time = 10 ms + 0.5 ms of dead time)
    time_val = [event * 10.5 for event in range(nb_mes)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, nb_mes))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_mes):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )
    return time_val, raw_strip_resp


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def calc_mean_val_plateau(steps, time_val, start_times, n_points, strip_resp):
    """function calculating the mean value of each step of step phantom for each strip.
    The start time calc of each step and nb of points used for averaging is given in
    argument."""

    mean_resp = np.zeros((len(strip_resp), len(steps)))
    std_resp = np.zeros((len(strip_resp), len(steps)))

    for iresp, resp in enumerate(strip_resp):
        for istep, step in enumerate(steps):
            start_event_index = find_closest_index(time_val, start_times[istep])
            last_event_index = int(start_event_index + n_points)
            mean_resp[iresp, istep] = np.mean(resp[start_event_index:last_event_index])
            std_resp[iresp, istep] = np.std(resp[start_event_index:last_event_index])

    return mean_resp, std_resp


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l, column=start_col + i, value=data_to_fill[i])
    wb_res.save(excel_path)


## DATA TO FILL #########
name_mes_file = r"\zData_150V_150ubeam_24p8v0_40step8_0cmvitesse10.bin"
file_mes_path = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\mesures_carac_det_8-06-2023\step_phantom\step_0_8cm\\zData_150V_150ubeam_24p8v0_40step8_0cmvitesse10.bin"
param_res_file = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"
ws_name = "resp_ESRF_microfx"
beam_type = "ESRF"
step_thicknesses = [0, 1, 2, 3, 4, 6]
fill_excel_bool = False


# Gets parameters values for analysis
df_param = pd.read_excel(param_res_file, sheet_name="param_analyse", skiprows=2)
param = df_param[beam_type].to_numpy()

# Nb of steps in step phantom
nb_of_steps = int(param[0])

# Number of points for noize calculation
nb_points_noize = int(param[1])

# Start time for normalization value calculation
n_points = param[3]

# Start time values used for averaging of each phantom steps
start_times = [time_val for time_val in param[4 : 4 + nb_of_steps]]
nb_points_mean_plateau_calc = int(param[10])

dt = np.dtype("uint16")
f = open(file_mes_path, "rb")
data = f.read()
zdata = np.frombuffer(data, dt)
time_val, raw_strip_resp = read_bin_file(zdata)


strip_resp, strip_std = calc_mean_val_plateau(
    step_thicknesses, time_val, start_times, n_points, raw_strip_resp
)

if fill_excel_bool:
    # Fill excel file with strip responses and uncertainties
    for iresp, resp in enumerate(strip_resp):
        fill_excel(
            param_res_file,
            ws_name,
            resp,
            iresp + 2,
            2,
        )

    for istd, std in enumerate(strip_std):
        fill_excel(
            param_res_file,
            "std",
            std,
            istd + 2,
            2,
        )
