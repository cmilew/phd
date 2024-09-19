import matplotlib.pyplot as plt
import numpy as np
import sys
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score
import matplotlib.ticker as ticker


def find_closest_index(list_of_values, target_value):
    return min(
        range(len(list_of_values)), key=lambda x: abs(list_of_values[x] - target_value)
    )


def exponential_model(x, a, b):
    return a * np.exp(-b * x)


def cut_noize_and_normal(raw_strip_resp, strip_noize, normalization_value, factor):
    clean_data_fn = (
        lambda x: (x - strip_noize) / (normalization_value - strip_noize) * factor
    )
    vectorized_clean_data_fn = np.vectorize(clean_data_fn)
    normal_strip_resp = vectorized_clean_data_fn(raw_strip_resp)
    return normal_strip_resp


def calc_mean_val_plateau(
    time_values, start_time_value, nb_of_averaged_points, strip_resp
):
    """function calculating the mean value of strip responses over a range of number of
    time values given in argument (with starting point), normally on a plateau of the
    step phantom. Returns an array of mean values for each strip."""

    start_event_index = find_closest_index(time_values, start_time_value)
    last_event_index = int(start_event_index + nb_of_averaged_points)
    mean_values_plateau = np.zeros(len(strip_resp))
    std_val_plateau = np.zeros(len(strip_resp))
    for strip in range(0, len(strip_resp)):
        mean_values_plateau[strip] = np.mean(
            strip_resp[strip, start_event_index:last_event_index]
        )
        std_val_plateau[strip] = np.std(
            strip_resp[strip, start_event_index:last_event_index]
        )
    return mean_values_plateau, std_val_plateau


def fill_mu_in_excel(ws_mu, strip_lists, mu_lists, start_l, start_col):
    for region in range(len(strip_lists)):
        for strip in range(len(strip_lists[region])):
            ws_mu.cell(
                row=start_l + strip, column=start_col, value=strip_lists[region][strip]
            )
            ws_mu.cell(
                row=start_l + strip, column=start_col + 1, value=mu_lists[region][strip]
            )
        wb_res.save(param_res_file)
        start_col += 3


def read_bin_file():
    # correspondence of QDC number and strip number file
    correspondence_table_file = r"C:\Users\milewski\Desktop\these\mesures\analyse_data\codes_python\150_voies\add_piste.txt"
    pf = open(correspondence_table_file, "r")
    correspondence_table = pf.readlines()

    # number of measurements
    nb_mes = np.size(zdata) // 309

    # time conversion (integration time = 10 ms + 0.5 ms of dead time)
    time_values = [event * 10.5 for event in range(nb_mes)]

    # strips responses matrix (line = strips, columns = strip responses)
    raw_strip_resp = np.zeros((153, nb_mes))

    # 17 first strips on the missing diamond => 0 response
    for strip_num in range(18, 153):
        corresponding_QDC_num = int(correspondence_table[strip_num])
        for event in range(nb_mes):
            raw_strip_resp[strip_num, event] = np.uint32(
                ((zdata[3 + corresponding_QDC_num * 2 + event * 309]) << 16)
                + (zdata[4 + corresponding_QDC_num * 2 + event * 309])
                >> 6
            )
    return time_values, raw_strip_resp


##### DATA TO FILL #######
beam_types = ["AlCu 0-1-2-2.5", "AlCu 0-3-4-5"]
ws_mu_name = "mu_AlCu"
line_width = 2
name_mes_files = [
    r"\4T_Al_Cu_step_fant_1_2_2.5_cm_VREF_UC_0_sans_colli_MRT.bin",
    r"\4T_Al_Cu_step_fant_3_4_5_cm_VREF_UC_10000_sans_colli_MRT.bin",
]
file_mes_path = r"C:\Users\milewski\Desktop\these\mission_australie\mesures\136_voies\20-03-2024\fantome_escalier\AlCu\MRT"
step_thicknesses = [0, 1, 2.5, 3, 4, 5]
###########################

# measurement files
mes_files = [file_mes_path + name_mes_files[i] for i in range(len(name_mes_files))]

# parameters and results file
param_res_file = r"C:\Users\milewski\Desktop\these\papiers\caracterisation_detecteur_153_voies\step_phantom\param_et_resultats.xlsx"
wb_res = load_workbook(param_res_file)
ws_mu = wb_res[ws_mu_name]

# mean and std response matrixes (line = strips, column = mean/std value for each step)
mean_strip_resp = []
std_mean_strip_resp = []
print(f"mean strip resp deb {mean_strip_resp}")
# loop filling previous lists with measured values
for i in range(len(mes_files)):
    dt = np.dtype("uint16")
    f = open(mes_files[i], "rb")
    data = f.read()
    zdata = np.frombuffer(data, dt)

    # gets parameters values for analysis
    beam_type = beam_types[i]
    df_param = pd.read_excel(param_res_file, sheet_name="param_analyse", skiprows=2)
    param = df_param[beam_type].to_numpy()

    # nb of steps in step phantom
    nb_of_steps = int(param[0])

    # nb of points for noize calculation
    nb_points_noize = int(param[1])

    # start time for normalization value calculation
    start_time_normal_calc = param[2]
    nb_points_normal_calc = param[3]

    # start time values used for averaging of each phantom steps
    start_time_values_each_thicknesses = [
        time_val for time_val in param[4 : 4 + nb_of_steps]
    ]
    nb_points_mean_plateau_calc = int(param[10])

    # get measurement values and calculate normalization value on step 0 plateau
    time_values, raw_strip_resp = read_bin_file()
    strip_normal_val, std_normal_values = calc_mean_val_plateau(
        time_values,
        start_time_normal_calc,
        nb_points_normal_calc,
        raw_strip_resp,
    )

    # noize calculation
    strip_noizes = [
        np.mean(raw_strip_resp[strip, 0:nb_points_noize]) for strip in range(18, 153)
    ]
    strip_noizes = np.concatenate((np.zeros(18), strip_noizes))

    # noize cut and normalization
    normal_strip_resp = np.zeros((153, len(time_values)))
    for strip in range(18, 153):
        normal_strip_resp[strip, :] = cut_noize_and_normal(
            raw_strip_resp[strip],
            strip_noizes[strip],
            strip_normal_val[strip],
            1,
        )

    for step in range(nb_of_steps):
        # if steps 0-3-4-5 file: skips step 0
        if "3_4_5" in mes_files[i] and step == 0:
            pass
        else:
            mean_val, std_val = calc_mean_val_plateau(
                time_values,
                start_time_values_each_thicknesses[step],
                nb_points_mean_plateau_calc,
                normal_strip_resp,
            )
            mean_strip_resp.append(mean_val)
            std_mean_strip_resp.append(std_val)


# Line and column for excel file writing
start_l_res = 4
start_col_res = 3

# exponential fit to retrieve linear att coeff (mu)
mu_values = np.zeros(153)
print(f"muvalue[18] {mu_values[18]}")
print(f"mean_strip_resp[18] {mean_strip_resp[0][18]}")
for strip in range(18, 153):
    popt, pcov = curve_fit(
        exponential_model,
        step_thicknesses,
        [mean_strip_resp[step][strip] for step in range(len(step_thicknesses))],
        p0=(1, 1),
    )
    a_opt, b_opt = popt
    mu_values[strip] = b_opt
print(f"mean_strip_resp[18] {mean_strip_resp[0][18]}")
print(f"muvalue[18] {mu_values[18]}")

# # # fills excel result file
# for strip_num, mu in enumerate(mu_values[18:153]):
#     strip_cell_val = ws_mu.cell(row=start_l_res, column=start_col_res).value
#     mu_cell_val = ws_mu.cell(row=start_l_res, column=start_col_res + 1).value
#     if strip_cell_val is not None:
#         ws_mu.cell(row=start_l_res, column=start_col_res).value = ""
#     if mu_cell_val is not None:
#         ws_mu.cell(row=start_l_res, column=start_col_res + 1).value = ""

#     ws_mu.cell(row=start_l_res, column=start_col_res).value = strip_num + 18
#     ws_mu.cell(row=start_l_res, column=start_col_res + 1).value = mu
#     start_l_res += 1
# print(f"muvalue[18] {mu_values[18]}")
# wb_res.save(param_res_file)

# sys.exit()
# plot mu values
# plt.scatter([strip for strip in range(18, 153)], strip_normal_val[18:153])
plt.scatter([strip for strip in range(18, 153)], mu_values[18:153])
plt.title("Mu distribution ANSTO poly Al/Cu")
plt.xlabel("Strip number")
plt.ylabel("Mu (cm-1)")
plt.show()


# fill_mu_in_excel(
#     ws_mu,
#     [mb_strips, interb_strips, off_p_strips],
#     [mb_mu, interb_mu, off_p_mu],
#     start_l_res,
#     start_col_res,
# )
# fill_mu_in_excel(ws_mu, interb_strips, interb_mu, start_l_res, start_col_res + 3)
# fill_mu_in_excel(ws_mu, off_p_strips, off_p_mu, start_l_res, start_col_res + 6)

# plot_hist_gaussian_fit(mb_mu, "microbeam strips")
# plot_hist_gaussian_fit(interb_mu, "interbeam strips")
# # print(f"off_p_mu {off_p_mu}")
# off_p_mu_cut = [off_p_mu[i] for i in range(len(off_p_mu)) if off_p_mu[i] > 0.154]
# print(f"off p mu {off_p_mu}")
# print(f"off p mu cut {off_p_mu_cut}")
# print(f"len off p mu cut {len(off_p_mu_cut)}")
# plt.hist(off_p_mu, bins=10, color="#173896", edgecolor="white")
# plt.title("Mass attenuation coeff off pitch strips")
# plt.xlabel("Mass attenuation coefficient (cm-1)")
# plt.ylabel("Number of strips")
# plt.show()
# plot_hist_gaussian_fit(off_p_mu_cut, "off pitch strips")
