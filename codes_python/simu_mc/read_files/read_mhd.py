import os
import SimpleITK as sitk
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np

# path_edep_uncertain = os.path.join(
#     os.path.dirname(__file__),
#     f"output/{slab_thickness}cm_rw3_strip_{strip_number}-edep-uncertainty.mhd",
# )
# gets dose actor = img
# edep = sitk.ReadImage(path_edep)
# # edep_uncertain = sitk.ReadImage(path_edep_uncertain)

# # converts dose actor img in numpy array
# edep_array = sitk.GetArrayFromImage(edep)
# edep_unertain_array = sitk.GetArrayFromImage(edep_uncertain)


# Charger le fichier .mhd
def load_mhd_image(file_path):
    """Load 2D image from file .mhd and convert it in numpy array."""
    image = sitk.ReadImage(file_path)  # Charger le fichier
    array = sitk.GetArrayFromImage(image)  # Convertir en tableau numpy
    return array[0, :, :]  # Extraire la seule coupe disponible (2D)


# Afficher l'image avec un nuancier de couleur
def display_image_with_colormap(image_array, colormap="viridis"):
    """Display a 2D image with a colormap."""

    plt.figure(figsize=(10, 8))
    plt.imshow(image_array, cmap=colormap)
    plt.colorbar()
    plt.axis("off")
    plt.show()


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l + i, column=start_col, value=data_to_fill[i])
    wb_res.save(excel_path)


# DATA TO FILL IN ##########
slab_thickness = 0.0
slab_material = "SoldHE"

###########################

# get mhd file path
script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
relative_path = f"output\{slab_thickness}mm_{slab_material}_detector_edep.mhd"
path_edep = os.path.join(script_dir, relative_path)
excel_path = os.path.join(script_dir, r"read_files\test_dose_actor.xlsx")

# load and convert image in np array
edep_array = load_mhd_image(path_edep)

# plt edep in function of x and y
# length_display = len(edep_array[0])
print(edep_array.shape)
length_display = 85
pos_x = range(length_display)
pos_y = np.ones(length_display)
plt.scatter(pos_x, pos_y, s=25, marker="s", c=edep_array[0][2046:2131], cmap="viridis")
cbar = plt.colorbar()
# cbar.set_label("Energy (MeV)", rotation=270, labelpad=15)
plt.ylim(0.999, 1.001)
# plt.xlim(-18, 18)
# plt.xlabel("x position (mm)")
# plt.ylabel("y position (mm)")
# plt.title(phsp_name)
plt.show()

fill_excel(excel_path, "Feuil1", edep_array[0], 2, 3)
