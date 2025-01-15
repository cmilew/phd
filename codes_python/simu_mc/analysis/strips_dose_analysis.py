from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import sys
from decimal import Decimal


def get_excel_data(excel_file, ws_name, start_l, start_col, len_data):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(excel_file)
    ws = wb_res[ws_name]
    excel_data = [
        ws.cell(row=start_l + i, column=start_col).value for i in range(len_data)
    ]
    return np.asarray(excel_data)


def generate_shifted_matrix(input_array):
    """Function generating a matrix where the given array is shifted at each row and
    missing values are filled with zeros"""
    n = len(input_array)
    result_matrix = np.zeros((n, n))
    if n % 2 == 0:
        center_index = n // 2 - 1
        shift = 1
    else:
        center_index = n // 2
        shift = 0
    index_result_matrix = 0

    for i in range(center_index + 1):
        result_matrix[index_result_matrix][: center_index + i + 1 + shift] = (
            input_array[center_index - i :]
        )
        index_result_matrix += 1
    for i in range(center_index + shift):
        result_matrix[index_result_matrix][i + 1 :] = input_array[0 : n - i - 1]
        index_result_matrix += 1

    return result_matrix


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l + i, column=start_col, value=data_to_fill[i])
    wb_res.save(excel_path)


# DATA TO FILL IN ######################
excel_file = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\simulations_MC\results_step_phantom.xlsx"
ws_name = "edep_136_strips"
n_particules = 100_000_000
start_l = 8
start_c = 3
n_strip_col = 2
n_strips = 136
central_strip = 68
log_scale = False
fill_excel_bool = False
########################################

# gets simulation data from excel file
strip_num = get_excel_data(excel_file, ws_name, start_l, n_strip_col, n_strips)
strip_edep = get_excel_data(excel_file, ws_name, start_l, start_c, n_strips)
strip_edep = strip_edep / n_particules * 1e6  # eV
strip_edep_rel_uncertain = get_excel_data(
    excel_file, ws_name, start_l, start_c + 1, n_strips
)


# calculates square of uncertainties for error propagation
strip_edep_uncertain = strip_edep * strip_edep_rel_uncertain
square_uncertain = np.square(strip_edep_uncertain)


# bar plot of 1 mb centered on central strip
plt.bar(strip_num, strip_edep, label="Données")
plt.errorbar(
    strip_num,
    strip_edep,
    ls="none",
    yerr=strip_edep_uncertain,
    capsize=2,
    label="Données",
    color="k",
)
plt.xlabel("Strip number", fontsize=15)
plt.xticks(np.arange(1, n_strips + 3, 3))
plt.ylabel("Edep / photon [eV]", fontsize=15)
plt.title(
    f"Simulations MC 1 microbeam {Decimal(n_particules):.0E} particles", fontsize=15
)
if log_scale:
    plt.yscale("log")
plt.xticks(rotation=45)
plt.show()

# sum strips edep
shifted_edep_matrix = generate_shifted_matrix(strip_edep)
shifted_uncertain_matrix = generate_shifted_matrix(square_uncertain)

# only keeps one line out of two of shifted matrix because mb every 2 strips
shifted_edep_matrix = shifted_edep_matrix[::2]
shifted_uncertain_matrix = shifted_uncertain_matrix[::2]

sum_edep = np.sum(shifted_edep_matrix, axis=0)
sum_square_uncertain = np.sum(shifted_uncertain_matrix, axis=0)
sum_uncertain = np.sqrt(sum_square_uncertain)

if fill_excel_bool:
    fill_excel(excel_file, "sum_136_strips", sum_edep, start_l, 15)
    fill_excel(excel_file, "sum_136_strips", sum_uncertain, start_l, 16)

# bar plot of mb matrix
plt.bar(strip_num, sum_edep, label="Données")
plt.errorbar(
    strip_num,
    sum_edep,
    ls="none",
    yerr=sum_uncertain,
    capsize=2,
    label="Données",
    color="k",
)
plt.xlabel("Strip number", fontsize=15)
plt.xticks(np.arange(1, n_strips + 3, 3))
plt.ylabel("Edep / photon [eV]", fontsize=15)
plt.title(
    "Sum of edep distribution of 1 microbeam MC simulation shifted every 2 strips",
    fontsize=15,
)
plt.xticks(rotation=45)
if log_scale:
    plt.yscale("log")
plt.show()

sys.exit()
# In beam/inter beam strip ratio
sum_edep_in_beam = sum_edep[::2]
sum_uncertain_in_beam = sum_uncertain[::2]
sum_edep_interb = sum_edep[1::2]
sum_uncertain_interb = sum_uncertain[1::2]

# Calculates ratio and propagates uncertainties
ratio = sum_edep_in_beam / sum_edep_interb
ratio_uncertain = ratio * np.sqrt(
    np.square(sum_uncertain_in_beam / sum_edep_in_beam)
    + np.square(sum_uncertain_interb / sum_edep_interb)
)

# remove last element because aberrant
# ratio = ratio[:-1]

# scatter plot of ratio
plt.scatter(range(len(ratio)), ratio)
plt.errorbar(
    range(len(ratio)),
    ratio,
    ls="none",
    yerr=ratio_uncertain,
    capsize=0,
    label="Données",
    color="k",
)
plt.xlabel("In beam strip number", fontsize=15)
plt.ylabel("In beam / inter beam ratio", fontsize=15)
plt.title(
    "In beam / inter beam ratio of sum edep distribution",
    fontsize=15,
)
plt.xticks(rotation=45)
plt.xticks(np.arange(1, 68, 2))
plt.show()
