from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import pandas as pd
import sys


def get_excel_data(data_file, ws_name, start_l, start_col, n_l, n_c):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(data_file)
    ws = wb_res[ws_name]
    excel_data = []
    for line in range(n_l):
        for col in range(n_c):
            excel_data.append(ws.cell(row=start_l + line, column=start_col + col).value)
    return np.asarray(excel_data)


def exponential_model(x, a, b):
    return a * np.exp(-b * x)


def calc_mu(step_thickness, edep):
    popt, pcov = curve_fit(
        exponential_model,
        step_thickness,
        edep,
        p0=(1, 1),
    )
    a_opt, b_opt = popt
    mu = b_opt

    # uncertainties on exponential fit
    perr = np.sqrt(np.diag(pcov))
    a_err, b_err = perr
    mu_uncertain = b_err

    return mu, mu_uncertain


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l + i * 2, column=start_col, value=data_to_fill[i])
    wb_res.save(excel_path)


def generate_normal_matrix(matrix, matrix_uncertain):
    """function normalizing every line of a matrix by the first element of the line and
    propagates uncertainties"""

    if matrix.shape != matrix_uncertain.shape:
        raise ValueError("matrix and matrix_uncertain must have the same shape")
    if matrix.ndim != 2:
        raise ValueError("matrix and matrix_uncertain must be 2D")

    normal_matrix = matrix / matrix[:, :1]
    normal_matrix_uncertain = normal_matrix * np.sqrt(
        (matrix_uncertain[:, :1] / matrix[:, :1]) ** 2
        + (matrix_uncertain / matrix) ** 2
    )
    return normal_matrix, normal_matrix_uncertain


def get_fit_expo_data(x_data, y_data):
    popt, pcov = curve_fit(
        exponential_model,
        x_data,
        y_data,
        p0=(1, 1),
    )
    a_opt, b_opt = popt
    mu = b_opt
    perr = np.sqrt(np.diag(pcov))
    a_err, b_err = perr
    mu_uncertain = b_err

    x_fit = np.linspace(min(step_thickness), max(step_thickness), len(step_thickness))
    y_fit = exponential_model(x_fit, *popt)

    return mu, mu_uncertain, x_fit, y_fit


# DATA TO FILL IN ######################
n_strips = 136
step_thickness = np.array([0, 1, 2, 3, 4, 6])
step_phant_mat = "RW3"
data_file = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\simulations_MC\results_step_phantom.xlsx"
ws_name = "ESRF_poly"
start_l_mc = 33
start_c_mc = 31
start_l_mes = 33
start_c_mes = 30
mu_NIST = 0.1616615
fill_excel_bool = False
fontsize_val = 20
# ########################################

# get edep of strips obtained from MC simu from excel file
df = pd.read_excel(data_file, sheet_name=ws_name)
subset_df = df.iloc[8 : 8 + n_strips, 2 : 2 + len(step_thickness)]
subset_df_uncertain = df.iloc[
    8 : 8 + n_strips,
    9 : 9 + len(step_thickness) * 2,
]
all_edep_mc = subset_df.to_numpy()
all_uncertain_mc = subset_df_uncertain.to_numpy()

all_edep_mc = all_edep_mc.astype(float)
all_uncertain_mc = all_uncertain_mc.astype(float)

# retrieves only statistical uncertainties
stat_uncertain_mc = all_uncertain_mc[:, 0::2]

# retrieves edep only for the strips facing microbeams and make sure values are float
edep_mc = np.array(all_edep_mc[1::2], dtype=float)
uncertain_mc = np.array(stat_uncertain_mc[1::2], dtype=float)

# calc mu for these strips
strip_mu_mc = []
uncertain_mu_fit_mc = []
for strip_edep in edep_mc:
    mu, uncertain_mu = calc_mu(step_thickness, strip_edep)
    strip_mu_mc.append(mu)
    uncertain_mu_fit_mc.append(uncertain_mu)
if fill_excel_bool:
    fill_excel(data_file, "test", strip_mu_mc, 2, 2)
    fill_excel(data_file, "test", uncertain_mu_fit_mc, 2, 3)


# get measurements and mc simul of strip 86 (= center strip of 8 diamonds) for plot
edep_mes_s86 = get_excel_data(data_file, ws_name, 10, 37, len(step_thickness), 1)
uncertain_mes_s86 = get_excel_data(data_file, ws_name, 19, 37, len(step_thickness), 1)

edep_mc_s86 = get_excel_data(data_file, ws_name, 10, 38, len(step_thickness), 1)
uncertain_mc_s86 = get_excel_data(data_file, ws_name, 19, 38, len(step_thickness), 1)


# # theoratical curve
# x_theo = np.arange(0, 5, 0.05, dtype=float)
# y_theo = np.exp(-mu_NIST * x_theo)

# # fit exponential model to mean of in beam sum edep
# mean_mu_mc, mean_mu_uncertain_mc, x_fit_mc, y_fit_mc = get_fit_expo_data(
#     step_thickness, in_beam_mean_mc
# )
# mu_mes, mu_mes_uncertain, x_fit_mes, y_fit_mes = get_fit_expo_data(step_thickness, mes)

# plt.plot(x_theo, y_theo, label="theory (NIST)", linestyle="--", color="black")

# strip 86 = 42 in strip_mu_mc because takes only microbeam lines
mu_mc_s86 = strip_mu_mc[42]
mu_uncertain_mc_s86 = uncertain_mu_fit_mc[42]

mu_mes_s86, uncertain_mu_mes_s86 = calc_mu(step_thickness, edep_mes_s86)

plt.errorbar(
    step_thickness,
    edep_mes_s86,
    yerr=uncertain_mes_s86,
    fmt="o",
    capsize=5,
    color="blue",
    label="experimental",
)
plt.errorbar(
    step_thickness,
    edep_mc_s86,
    yerr=uncertain_mc_s86,
    fmt="^",
    capsize=5,
    color="orange",
    label="MC simulations",
)
plt.plot(
    [],
    [],
    color="white",
    label=f"$\mu$ MC = {mu_mc_s86:.3f} $\pm$ {round(mu_uncertain_mc_s86, 3)} cm\u207b\u00b9",
)
plt.plot(
    [],
    [],
    color="white",
    label=f"$\mu$ exp = {mu_mes_s86:.3f} $\pm$ {round(uncertain_mu_mes_s86, 3)} cm\u207b\u00b9",
)

x_label = step_phant_mat + " thickness (cm)"
plt.xlabel(x_label, fontsize=fontsize_val)
plt.ylabel("Response / Response step 0", fontsize=fontsize_val)
plt.tick_params(axis="x", labelsize=fontsize_val)
plt.tick_params(axis="y", labelsize=fontsize_val)
plt.legend(fontsize=fontsize_val)
plt.show()
