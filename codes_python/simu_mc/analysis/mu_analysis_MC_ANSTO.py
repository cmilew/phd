from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import pandas as pd
import sys


def get_excel_data(data_file, ws_name, start_l, start_col, n_l, n_c):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(data_file)
    ws = wb_res[ws_name]
    excel_data = []
    for line in range(n_l):
        for col in range(n_c):
            excel_data.append(ws.cell(row=start_l + line, column=start_col + col).value)
    return np.asarray(excel_data)


def exponential_model(x, a, b):
    return a * np.exp(-b * x)


def calc_mu(step_thickness, edep, uncertain_edep):
    # convert in np.array type dfloat to prevent errors
    step_thickness = np.array(step_thickness, dtype=float)
    edep = np.array(edep, dtype=float)
    uncertain_edep = np.array(uncertain_edep, dtype=float)

    # Replace uncertainties = 0 by small value to prevent errors
    uncertain_edep[uncertain_edep == 0] = 1e-6

    # initial parameteres estimation (to help convergence)
    p0 = [edep[0], 0.01]  # a = première valeur de edep, b ≈ petite valeur positive
    popt, pcov = curve_fit(
        exponential_model,
        step_thickness,
        edep,
        sigma=uncertain_edep,
        absolute_sigma=True,
        p0=p0,
    )
    a_opt, b_opt = popt
    mu = b_opt

    # uncertainties on exponential fit
    perr = np.sqrt(np.diag(pcov))
    mu_uncertain = perr[1] if perr[1] != np.inf else np.nan  # handle +/- inf case

    return mu, mu_uncertain


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l + i, column=start_col, value=data_to_fill[i])
    wb_res.save(excel_path)


def generate_normal_matrix(matrix, matrix_uncertain):
    """function normalizing every line of a matrix by the first element of the line and
    propagates uncertainties"""

    if matrix.shape != matrix_uncertain.shape:
        raise ValueError("matrix and matrix_uncertain must have the same shape")
    if matrix.ndim != 2:
        raise ValueError("matrix and matrix_uncertain must be 2D")

    normal_matrix = matrix / matrix[:, :1]
    normal_matrix_uncertain = normal_matrix * np.sqrt(
        (matrix_uncertain[:, :1] / matrix[:, :1]) ** 2
        + (matrix_uncertain / matrix) ** 2
    )
    return normal_matrix, normal_matrix_uncertain


# DATA TO FILL IN ######################
n_strips = 136
step_thickness = np.array([0, 1, 2, 2.5, 3, 4, 5])
step_phant_mat = "SolidHE"
data_file = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\simulations_MC\results\res_ANSTO_poly_CuCu_step_phant_MC_simu.xlsx"
ws_name = "ANSTO_poly_CuCu_MC"
fill_excel_bool = True
fontsize_val = 20
plot_res_bool = False
mu_mc = 0.179
uncertain_mu_mc = 0.002
mu_mes = 0.176
mu_NIST = 0.177
uncertain_mu_mes = 0.001
# ########################################

# get edep of strips obtained from MC simu from excel file
df = pd.read_excel(data_file, sheet_name=ws_name)
subset_df = df.iloc[8 : 8 + n_strips, 24 : 24 + len(step_thickness)]
subset_df_uncertain = df.iloc[
    8 : 8 + n_strips,
    32 : 32 + len(step_thickness),
]
edep_mc = subset_df.to_numpy()
uncertain_mc = subset_df_uncertain.to_numpy()

# print(f"edep_mc[0, :] = {edep_mc[0, :]}")
# print(f"uncertain_mc[0, :] = {uncertain_mc[0, :]}")
# print(f"edep_mc.shape = {edep_mc.shape}")
# print(f"uncertain_mc.shape = {uncertain_mc.shape}")
# print(f"edep_mc[-1] = {edep_mc[-1]}")
# print(f"uncertain_mc[-1] = {uncertain_mc[-1]}")
# sys.exit()

# edep_mc = edep_mc.astype(np.float32)
# uncertain_mc = uncertain_mc.astype(np.float32)

# print(edep_mc[:, 1])
# # print(uncertain_mc.shape)
# sys.exit()

# calc mu for every strips
strip_mu_mc = []
uncertain_mu_fit_mc = []
for strip in range(n_strips):
    mu, uncertain_mu = calc_mu(step_thickness, edep_mc[strip], uncertain_mc[strip])
    strip_mu_mc.append(mu)
    uncertain_mu_fit_mc.append(uncertain_mu)
if fill_excel_bool:
    fill_excel(data_file, "test", strip_mu_mc, 2, 2)
    fill_excel(data_file, "test", uncertain_mu_fit_mc, 2, 3)


# # get measurements edep for high pvdr strip facing mb (mean of value) and strip 25 for MC
# edep_mes_high_pvdr = get_excel_data(data_file, ws_name, 10, 39, len(step_thickness), 1)
# uncertain_mes_high_pvdr = get_excel_data(
#     data_file, ws_name, 19, 39, len(step_thickness), 1
# )

# edep_mc_s23 = get_excel_data(data_file, ws_name, 10, 40, len(step_thickness), 1)
# uncertain_mc_s23 = get_excel_data(data_file, ws_name, 19, 40, len(step_thickness), 1)

# print(f"edep_mes_high_pvdr = {edep_mes_high_pvdr}")
# print(f"uncertain_mes_high_pvdr = {uncertain_mes_high_pvdr}")
# print(f"edep_mc_s23 = {edep_mc_s23}")
# print(f"uncertain_mc_s23 = {uncertain_mc_s23}")
# # print(f"edep_mc[24, :] = {edep_mc[24, :]}")
# sys.exit()

# Plot results


if plot_res_bool:
    # for theoratical curve
    x_theo = np.arange(0, 5, 0.05, dtype=float)
    y_theo = np.exp(-mu_NIST * x_theo)

    y_mc = np.exp(-mu_mc * x_theo)
    y_mes = np.exp(-mu_mes * x_theo)

    plt.errorbar(
        step_thickness,
        edep_mes_high_pvdr,
        yerr=uncertain_mes_high_pvdr,
        fmt="o",
        capsize=5,
        color="orange",
    )
    plt.errorbar(
        step_thickness,
        edep_mc_s23,
        yerr=uncertain_mc_s23,
        fmt="^",
        capsize=5,
        color="blue",
    )
    plt.plot(x_theo, y_mes, label="experimental", color="orange")
    plt.plot(x_theo, y_theo, label="theory (NIST)", linestyle="--", color="black")
    plt.plot(x_theo, y_mc, label="MC simulations", color="blue")

    plt.plot(
        [],
        [],
        color="white",
        label=f"$\mu$ exp = {mu_mes} $\pm$ {uncertain_mu_mes} cm\u207b\u00b9",
    )
    plt.plot(
        [],
        [],
        color="white",
        label=f"$\mu$ NIST = {mu_NIST:.3f} cm\u207b\u00b9",
    )
    plt.plot(
        [],
        [],
        color="white",
        label=f"$\mu$ MC = {mu_mc} $\pm$ {uncertain_mu_mc} cm\u207b\u00b9",
    )

    x_label = step_phant_mat + " thickness (cm)"
    plt.xlabel(x_label, fontsize=fontsize_val)
    plt.ylabel("Response / Response step 0", fontsize=fontsize_val)
    plt.tick_params(axis="x", labelsize=fontsize_val)
    plt.tick_params(axis="y", labelsize=fontsize_val)
    plt.legend(fontsize=fontsize_val)
    plt.show()
