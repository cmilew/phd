import matplotlib.pyplot as plt
import numpy as np
import sys
import pandas as pd
from openpyxl import load_workbook


def get_excel_data(data_file, ws_name, start_l, start_col, n_l, n_c):
    """gets normalization values from excel file calculated on step 0 of step phantom
    measurements"""
    wb_res = load_workbook(data_file)
    ws = wb_res[ws_name]
    excel_data = []
    for line in range(n_l):
        for col in range(n_c):
            excel_data.append(ws.cell(row=start_l + line, column=start_col + col).value)
    return np.asarray(excel_data)


#### DATA TO FILL #########
data_file = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\simulations_MC\results\res_ANSTO_poly_CuCu_step_phant_MC_simu.xlsx"
ws_name = "ANSTO_poly_CuCu_MC"
n_strips = 136
step_thickness = np.array([0, 1, 2, 2.5, 3, 4, 5])
mc_simu = True
fill_excel_bool = True

# get strip edep
df = pd.read_excel(data_file, sheet_name=ws_name)
subset_df = df.iloc[8 : 8 + n_strips, 1 : 1 + len(step_thickness)]
edep = subset_df.to_numpy()
edep = edep.astype(float)

# Uncertainties
if mc_simu:
    subset_df_uncertain = df.iloc[
        8 : 8 + n_strips,
        9 : 9 + len(step_thickness) * 2,
    ]
    all_uncertain = subset_df_uncertain.to_numpy()
    all_uncertain = all_uncertain.astype(float)
    # retrieves only statistical uncertainties (for MC simulations)
    uncertain = all_uncertain[:, 0::2]

else:
    subset_df_uncertain = df.iloc[
        8 : 8 + n_strips,
        8 : 8 + len(step_thickness),
    ]
    uncertain = subset_df_uncertain.to_numpy()
    uncertain = uncertain.astype(float)


# Normalize by value on first ste^p
normal_edep = edep / edep[:, [0]]
normal_uncertain = normal_edep * np.sqrt(
    (uncertain / edep) ** 2 + (uncertain[:, [0]] / edep[:, [0]]) ** 2
)

if fill_excel_bool:
    wb_res = load_workbook(data_file)
    ws = wb_res[ws_name]
    # Fill excel file with edep normalized and propagated uncertainties
    for iresp, resp in enumerate(normal_edep):
        for step in range(len(resp)):
            ws.cell(row=iresp + 10, column=step + 25, value=resp[step])
            ws.cell(
                row=iresp + 10, column=step + 33, value=normal_uncertain[iresp, step]
            )
    wb_res.save(data_file)
