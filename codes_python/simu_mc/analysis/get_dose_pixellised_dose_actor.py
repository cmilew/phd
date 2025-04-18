import os
import SimpleITK as sitk
import numpy as np
from openpyxl import load_workbook
import sys
import matplotlib.pyplot as plt


def get_strip_edep(edep_array, edep_uncertain_array):
    """Dose actor pixellised : 23 pixels/strip and 8 pixels/interstrip zones
    This function sum the dose in 23 adjacent pixels every 8 pixels starting by 8
    8 pixels (side )"""

    # mhd x y z => numpy z y x
    edep_x = edep_array[0][0]
    uncertain_x = edep_uncertain_array[0][0]
    strip_edep = []
    strip_uncertain = []
    n_strip_pix = 23
    n_interstrip_pix = 8
    total_pix = len(edep_x)
    pixel = 0

    while pixel < total_pix:
        if pixel + n_strip_pix <= total_pix:
            strip_edep.append(np.sum(edep_x[pixel : pixel + n_strip_pix]))
            strip_uncertain.append(
                np.sqrt(np.sum(uncertain_x[pixel : pixel + n_strip_pix] ** 2))
            )

        else:
            break
        pixel += n_strip_pix + n_interstrip_pix

    return np.array(strip_edep), np.array(strip_uncertain)


def fill_excel(excel_path, ws_name, data_to_fill, start_l, start_col):
    wb_res = load_workbook(excel_path)
    ws = wb_res[ws_name]
    for i in range(len(data_to_fill)):
        ws.cell(row=start_l + i, column=start_col, value=data_to_fill[i])
    wb_res.save(excel_path)


def plot_hist(val, bins, label_x, label_y, title):
    plt.hist(val, bins=bins, color="skyblue", edgecolor="black")
    plt.xlabel(label_x)
    plt.ylabel(label_y)
    plt.title(title)
    plt.grid(axis="y", linestyle="--", alpha=0.7)
    plt.show()


def plot_mean_diff(strip_edep):
    """function calculating the difference between strip edep and mean strip
    edep for microbeam and interbeam zones"""

    mean_edep_mb = np.mean(strip_edep[1::2])
    mean_diff_mb = strip_edep[1::2] - mean_edep_mb
    mean_edep_intermb = np.mean(strip_edep[0::2])
    mean_diff_intermb = strip_edep[0::2] - mean_edep_intermb

    label_x = "edep - " + r"$\bar{edep}$"
    plot_hist(
        mean_diff_mb, label_x, "f", "Histogramme écart à la moyenne zone microfaisceau"
    )
    plot_hist(
        mean_diff_intermb,
        label_x,
        "f",
        "Histogramme écart à la moyenne zone interfaisceau",
    )


### DATA TO FILL ######
slab_thickness = 10.0
slab_material = "SolidHE"
ws_name = "ANSTO_poly_CuCu_MC"
# ws_name = "test"
fill_excel_bool = True
plot_bool = False
mean_diff_bool = False
edep_col_to_fill = 3
uncertain_col_to_fill = 12
# synchrotron = "ESRF"
synchrotron = "ANSTO_poly_CuCu_MC"
excel_path = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\simulations_MC\results\res_ANSTO_poly_CuCu_step_phant_MC_simu.xlsx"

# simulation files
path_edep = os.path.join(
    os.path.dirname(__file__),
    rf"{ws_name}/{slab_thickness}mm_{slab_material}_detector_edep_tot.mhd",
)
path_edep_uncertain = os.path.join(
    os.path.dirname(__file__),
    rf"{ws_name}/{slab_thickness}mm_{slab_material}_detector_edep_uncertainty_tot.mhd",
)

# gets dose actor = img
edep = sitk.ReadImage(path_edep)
edep_rel_uncertain = sitk.ReadImage(path_edep_uncertain)

# converts dose actor img in numpy array
edep_array = sitk.GetArrayFromImage(edep)
edep_rel_uncertain_array = sitk.GetArrayFromImage(edep_rel_uncertain)
edep_uncertain = edep_array * edep_rel_uncertain_array / 100

# fill_excel(excel_path, ws_name, edep_array[0][0], 5, 4)
# sys.exit()


strip_edep, strip_uncertain = get_strip_edep(edep_array, edep_uncertain)

# calc 2 sigma for mb
mean_edep_mb = np.mean(strip_edep[1::2])
sigma = np.std(strip_edep[1::2])
minus_2sigma = mean_edep_mb - 2 * sigma
plus_2sigma = mean_edep_mb + 2 * sigma


if plot_bool:
    slab_thickness = int(slab_thickness)
    plt.bar(range(1, 137), strip_edep, capsize=3)
    # plt.axhline(mean_edep_mb, color="red", linestyle="--", label="mean microbeam")
    # plt.axhline(minus_2sigma, color="green", linestyle="--", label="2 sigma")
    # plt.axhline(plus_2sigma, color="green", linestyle="--")
    plt.xlabel("Strip number")
    plt.ylabel("Edep (MeV)")
    plt.title(f"Edep measured by strips {synchrotron}")
    plt.legend()
    plt.show()

if mean_diff_bool:
    plot_mean_diff(strip_edep)

if fill_excel_bool:
    strip_rel_uncertain = strip_uncertain / strip_edep * 100
    fill_excel(excel_path, ws_name, strip_edep, 10, edep_col_to_fill)
    fill_excel(excel_path, ws_name, strip_uncertain, 10, uncertain_col_to_fill)
    fill_excel(excel_path, ws_name, strip_rel_uncertain, 10, uncertain_col_to_fill + 1)
    print(f"Excel file {excel_path} filled")
