from openpyxl import load_workbook
import itertools

EXCEL_PATH = "C:/Users/milewski/Desktop/these/papiers/caracterisation_detecteur_153_voies/simulations_MC/ESRF_spectrum.xlsx"
EXCEL_SHEET = "Sheet1"
DATA_STARTING_AT = {"row": 7, "column": 1}
DATA_OFFSET = {"row": 0, "column": 8}
COLUMN_TITLES = ("Energy [keV]", "Flux 1 (agg)", "Flux 2 (agg)", "Flux 3 (agg)")

wb = load_workbook(EXCEL_PATH)
ws = wb[EXCEL_SHEET]

for icolumn in range(len(COLUMN_TITLES)):
    ws.cell(
        row=DATA_STARTING_AT["row"] + DATA_OFFSET["row"],
        column=DATA_STARTING_AT["column"] + DATA_OFFSET["column"] + icolumn,
        value=COLUMN_TITLES[icolumn],
    )

for padded_irow in range(DATA_STARTING_AT["row"] + 1, ws.max_row):
    irow = padded_irow - DATA_STARTING_AT["row"]

    if irow % 10 == 0:
        ws.cell(
            row=padded_irow + (irow // 10) - 1,
            column=DATA_STARTING_AT["column"] + DATA_OFFSET["column"],
            value=ws.cell(
                row=padded_irow,
                column=DATA_STARTING_AT["column"],
            ).value,
        )

    for icol in range(len(COLUMN_TITLES) - 1):
        prev_val = (
            ws.cell(
                row=padded_irow + (irow % 10),
                column=DATA_STARTING_AT["column"] + DATA_OFFSET["column"] + icol + 1,
            ).value
            or 0
        )

        add_val = (
            ws.cell(
                row=padded_irow,
                column=DATA_STARTING_AT["column"] + icol + 1,
            ).value
            or 0
        )

        print(
            f"irow: {(irow // 10)}, padded_irow: {padded_irow}, col {DATA_STARTING_AT['column'] + icol + 1}, prev_val: {prev_val}, add_val: {add_val}"
        )

        ws.cell(
            row=padded_irow + (irow // 10),
            column=DATA_STARTING_AT["column"] + DATA_OFFSET["column"] + icol + 1,
            value=prev_val + add_val,
        )

    irow += 1

wb.save(EXCEL_PATH + "_modified.xlsx")
