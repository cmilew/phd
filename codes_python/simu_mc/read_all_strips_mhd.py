import os
import SimpleITK as sitk
from openpyxl import load_workbook
import sys
import re

# DATA TO FILL IN ######################
rw3_thickness = 0
output_path = r"C:\Users\milewski\Desktop\these\phd\codes_python\simu_mc\run.yfauxwpb"
excel_path = r"C:\Users\milewski\OneDrive - Université Grenoble Alpes\these\papiers\caracterisation_detecteur_153_voies\simulations_MC\results_step_phantom.xlsx"
ws_name = "test_script_sum_mhd"
start_l = 6
########################################

wb_res = load_workbook(excel_path)
ws = wb_res[ws_name]

for filename in os.listdir(output_path):
    if filename.endswith(".mhd"):
        # gets strip number
        strip_number = re.search(r"(\d{1,3})\s*-edep", filename)
        # assert len(strip_number.group(1)) == 1, "strip number not found"
        strip_number = int(strip_number.group(1))

        # reads dose actor file
        file_path = os.path.join(output_path, filename)
        img = sitk.ReadImage(file_path)

        # converts img of dose actor in array (only one pixel)
        strip_value = sitk.GetArrayFromImage(img)
        if "uncertainty" in filename.lower():
            ws.cell(row=strip_number + start_l, column=6, value=strip_value[0][0][0])
        else:
            ws.cell(row=strip_number + start_l, column=5, value=strip_value[0][0][0])

wb_res.save(excel_path)
